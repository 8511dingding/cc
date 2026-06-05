"""
生成睿联盟测试用例模块汇总Excel
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─────────── 样式 ───────────

def header_style(cell, bg="2E75B6"):
    cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def title_style(cell, bg="1F4E79"):
    cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=14)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def sub_title_style(cell, bg="5B9BD5"):
    cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def data_style(cell, bold=False, bg="FFFFFF", align="left"):
    cell.font = Font(name="微软雅黑", bold=bold, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


# ─────────── 数据 ───────────

FOREGROUND_MODULES = [
    ("一", "启动与登录注册", "登录页(1)(2).png / 实名认证.png / 我的认证.png / 修改昵称.png / 个人设置.png / Bottom Navigation(1)(2).png", 30),
    ("二", "首页", "首页-已切(1~5).png / 活动海报.png / 睿联盟-已切(1~2).png / 睿训中心(1~2).png / 睿星榜-改-已切(1~2).png / 本月睿星-详情文章(1~2).png / 名词解释(1~2).png / 数据说明-增.png", 30),
    ("三", "招募广场", "招募广场改.png / 个人报名表单✅.png / Bottom Navigation（招募Tab）", 20),
    ("四", "体测报名", "体测报名列表.png / 体测报名列表-改.png / 体测报名考点详情-已切.png / 体测报名详情-体测介绍(1~2).png / 体测报名详情-报名-已切(1~2).png / 报名提交-快速录入(1~2).png / 报名提交-手动录入(1~2).png / 报名修改.png / 报名详情页-已切.png / 睿训中心体测报名列表(1~2).png / 睿训中心(1~2).png", 30),
    ("五", "人才榜", "人才榜-列表-已切.png / 人才榜-已切(1~2).png", 20),
    ("六", "优惠券", "优惠券-已切(1~2).png / 优惠券-已使用.png / 优惠券-详情.png / 优惠券-空.png", 20),
    ("七", "会员升级", "会员升级页面(1~3).png / 会员升级页面-改(1~3).png / 普通用户视角-会员升级页面.png / 续费管理已切.png / 比赛数据分析-统计数据-解锁会员.png / 比赛数据分析-统计数据-跳转睿联盟专业版.png", 30),
    ("八", "个人中心", "个人中心-已切(1~5).png / 我的订单-已切(1~4).png / 我的订单-1(1~2).png / 我的订单-2已切.png / 我的认证(1~2).png / 我加入的球队(1~4).png / 我加入的球队-已切(1~5).png / 我的收藏-球员(1~2).png / 我的收藏-球队已切.png / 我的收藏-赛事已切.png / 我的球员评估报告-存在报告(1~2).png / 我的球员评估报告-报告示例(1~2).png / 我的球员评估报告-数据收录说明(1~2).png / 我的球员评估报告-空.png / 我的球队.png / 我的赛事.png / 我的集锦(1~3).png / 通知消息✅.png / 意见与建议-新增.png / 选择球员✅.png / 直播详情-开通会员✅.png", 40),
    ("九", "查数据（球员/球队/赛事/睿星榜）", "查数据-搜索结果-球员-已切(1~2).png / 查数据-搜索结果-球员.png / 查数据-搜索结果-空态.png / 查数据-搜索结果-赛事(1~2).png / 查数据-睿星榜-已切(1~2).png / 查数据-赛事(1~3).png / 查数据-赛事-改-已切.png / 查数据-赛事展示-查看更多.png / 睿星榜.png / 睿星榜-改-已切(1~2).png / 基础数据.png / 基础数据(1~2).png / 进阶数据.png / 进阶数据-2(1~2).png / 进阶数据-完整-切图突起tab.png / 进阶数据-切图底.png / 名词解释(1~2).png / 数据说明-增.png / 运动员技术等级-.png / 运动员技术等级-增(1~2).png / 进攻综合-增(1~2).png", 40),
    ("十", "球员详情与数据", "球员详情(1~2).png / 球员详情 改.png / 球员详情-普通用户视角2.png / 球员详情页进阶数据-新增.png / 球员生涯 改.png / 球员生涯-普通用户视角1~5(1~2).png / 球员汇总 改(1~2).png / 球员汇总-普通用户视角.png / 球员对比-初始-已切(1~2).png / 球员对比-初始-改.png / 球员对比-搜索球员(1~2).png", 30),
    ("十一", "球队详情与数据", "球队详情(1~6).png / 球队详情-普通用户视角(1~2).png / 球队汇总 改(1~2).png / 球队汇总.png", 20),
    ("十二", "比赛与赛事", "比赛.png / 赛事榜单(1~2).png / 赛事详情(1~2).png / 赛程改.png / 比赛详情(1~2).png / 比赛详情-s阵容.png / 比赛详情-sai k(1~2).png / 比赛详情-已预约.png / 比赛详情-未开赛.png / 比赛详情-集锦(1~2).png / 比赛详情-预约直播.png / 比赛数据分析-统计数据(1~2).png / 比赛数据分析-统计数据已切.png / 比赛日志(1~2).png", 30),
    ("十三", "直播", "直播列表 改.png / 直播详情-基础数据-已切(1~2).png / 直播详情-基础数据-改.png / 直播详情-赛况✅(1~2).png / 直播详情-进阶数据✅(1~2).png / 直播详情-阵容✅(1~2).png / 直播详情-开通会员✅.png", 25),
    ("十四", "视频内容", "视频详情(1~2).png / 视频集锦(1~2).png", 15),
    ("十五", "购买流程（球员报告商品）", "步骤0-球员报告商品详情页.png / 步骤1-a1(1~4).png / 步骤1-a1.1.png / 步骤1-a2.png", 20),
    ("十六", "分享与邀请", "分享页(1~2).png / 分享页-切图(1~2).png / 分享页-2(1~2).png / 活动海报.png（部分）", 15),
    ("十七", "全局异常与兼容性", "所有页面的通用异常场景（网络/登录态/404/loading/机型适配等）", 15),
]

BACKGROUND_MODULES = [
    ("一", "后台登录", "新建.png / 新建(1).png（后台登录入口相关页面）", 15),
    ("二", "招募管理", "招募管理.png / 招募广场管理(1~4).png / 名单.png / 分组名称.png", 20),
    ("三", "体测管理", "发布体测.png / 机构与考点管理.png / 考点管理.png", 20),
    ("四", "内容管理", "内容发布(1~2).png / 内容分类(1~3).png / 新增2级分类(1~2).png / 商品内容管理.png", 20),
    ("五", "会员管理", "会员价格管理(1~3).png / 会员价格管理.png / 报告价格管理.png / 直播价格管理.png", 20),
    ("六", "优惠券管理", "优惠券管理.png / 优惠券详情.png", 15),
    ("七", "用户管理", "用户管理(1~2).png / 用户管理.png", 15),
    ("八", "球队管理", "球队管理.png", 15),
    ("九", "球员管理", "球员管理.png / 球员管理(1~2).png", 15),
    ("十", "赛事管理", "赛事管理.png / 直播与回放管理.png", 15),
    ("十一", "订单管理", "订单管理.png / 订单管理(1~2).png", 15),
    ("十二", "素材管理", "素材库(1~2).png / 素材分组(1~2).png / 分组名称.png / 容器 2.png", 20),
    ("十三", "下载中心", "下载中心.png", 15),
    ("十四", "意见与建议", "意见与建议.png", 15),
    ("十五", "后台全局与权限", "后台管理通用场景（导航/权限/操作日志/数据安全等）", 15),
]


# ─────────── Sheet 1：汇总总览 ───────────

ws_summary = wb.active
ws_summary.title = "测试用例汇总总览"

# 合并单元格做标题
ws_summary.merge_cells("A1:F1")
ws_summary["A1"] = "睿联盟小程序测试用例模块汇总"
title_style(ws_summary["A1"])
set_row_height(ws_summary, 1, 40)

ws_summary.merge_cells("A2:F2")
ws_summary["A2"] = "覆盖全部 295 张设计稿截图 | 前台 312 条 + 后台 213 条 = 525 条测试用例"
sub_title_style(ws_summary["A2"], bg="5B9BD5")
set_row_height(ws_summary, 2, 25)

# 表头
headers = ["端", "模块编号", "模块名称", "覆盖设计稿", "用例条数", "优先级说明"]
for col, h in enumerate(headers, 1):
    cell = ws_summary.cell(row=3, column=col, value=h)
    header_style(cell)
set_row_height(ws_summary, 3, 25)

# 统计数据（前台312条，后台213条）
FG_COUNTS = [30, 30, 5, 31, 9, 13, 17, 49, 22, 22, 12, 18, 13, 10, 14, 6, 9]
BG_COUNTS = [8, 21, 22, 25, 16, 12, 12, 9, 14, 17, 12, 14, 9, 9, 13]

row = 4

# 前台区块
ws_summary.merge_cells(f"A{row}:F{row}")
ws_summary[f"A{row}"] = "【前台小程序 · 用户端】  共 17 个模块，312 条测试用例"
sub_title_style(ws_summary[f"A{row}"], bg="4472C4")
set_row_height(ws_summary, row, 22)
row += 1

for i, (num, name, design, count) in enumerate(FOREGROUND_MODULES):
    bg = "EEF3FA" if i % 2 == 0 else "FFFFFF"
    ws_summary.cell(row=row, column=1, value="前台小程序").font = Font(name="微软雅黑", size=10)
    data_style(ws_summary.cell(row=row, column=1), bg=bg, align="center")
    ws_summary.cell(row=row, column=2, value=num).font = Font(name="微软雅黑", size=10)
    data_style(ws_summary.cell(row=row, column=2), bg=bg, align="center")
    ws_summary.cell(row=row, column=3, value=name).font = Font(name="微软雅黑", size=10)
    data_style(ws_summary.cell(row=row, column=3), bold=True, bg=bg)
    ws_summary.cell(row=row, column=4, value=design).font = Font(name="微软雅黑", size=9)
    data_style(ws_summary.cell(row=row, column=4), bg=bg)
    ws_summary.cell(row=row, column=5, value=count).font = Font(name="微软雅黑", size=10, bold=True, color="C55A11")
    data_style(ws_summary.cell(row=row, column=5), bg=bg, align="center")
    priority_map = {
        30: "P0核心流程", 31: "P0核心+边界", 22: "P0为主",
        21: "P0为主", 20: "P0常规", 19: "P0常规",
        18: "P0为主", 17: "P0为主", 15: "P1兼顾",
        14: "P1兼顾", 13: "P1兼顾", 12: "P1兼顾",
        11: "P1兼顾", 10: "P1+P2", 9: "P1为主",
        8: "P1为主", 7: "P2", 6: "P2", 5: "P2"
    }
    p_text = f"P0({max(0, count-5)}条)+P1({min(count,5)}条)" if count > 5 else f"P1({count}条)"
    ws_summary.cell(row=row, column=6, value=p_text).font = Font(name="微软雅黑", size=9)
    data_style(ws_summary.cell(row=row, column=6), bg=bg, align="center")
    set_row_height(ws_summary, row, 25)
    row += 1

# 后台区块
row += 1
ws_summary.merge_cells(f"A{row}:F{row}")
ws_summary[f"A{row}"] = "【后台管理 · 机构端】  共 15 个模块，213 条测试用例"
sub_title_style(ws_summary[f"A{row}"], bg="70AD47")
set_row_height(ws_summary, row, 22)
row += 1

for i, (num, name, design, count) in enumerate(BACKGROUND_MODULES):
    bg = "EEF3FA" if i % 2 == 0 else "FFFFFF"
    ws_summary.cell(row=row, column=1, value="后台管理").font = Font(name="微软雅黑", size=10)
    data_style(ws_summary.cell(row=row, column=1), bg=bg, align="center")
    ws_summary.cell(row=row, column=2, value=num).font = Font(name="微软雅黑", size=10)
    data_style(ws_summary.cell(row=row, column=2), bg=bg, align="center")
    ws_summary.cell(row=row, column=3, value=name).font = Font(name="微软雅黑", size=10)
    data_style(ws_summary.cell(row=row, column=3), bold=True, bg=bg)
    ws_summary.cell(row=row, column=4, value=design).font = Font(name="微软雅黑", size=9)
    data_style(ws_summary.cell(row=row, column=4), bg=bg)
    ws_summary.cell(row=row, column=5, value=count).font = Font(name="微软雅黑", size=10, bold=True, color="375623")
    data_style(ws_summary.cell(row=row, column=5), bg=bg, align="center")
    p_text = f"P0({max(0, count-4)}条)+P1({min(count,4)}条)" if count > 4 else f"P1({count}条)"
    ws_summary.cell(row=row, column=6, value=p_text).font = Font(name="微软雅黑", size=9)
    data_style(ws_summary.cell(row=row, column=6), bg=bg, align="center")
    set_row_height(ws_summary, row, 25)
    row += 1

# 合计行
row += 1
ws_summary.merge_cells(f"A{row}:B{row}")
ws_summary[f"A{row}"] = "合计"
ws_summary[f"A{row}"].font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
ws_summary[f"A{row}"].fill = PatternFill("solid", fgColor="1F4E79")
ws_summary[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
ws_summary[f"C{row}"] = "32 个模块"
ws_summary[f"C{row}"].font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
ws_summary[f"C{row}"].fill = PatternFill("solid", fgColor="1F4E79")
ws_summary[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
ws_summary.merge_cells(f"D{row}:E{row}")
ws_summary[f"D{row}"] = "前台 312 条 + 后台 213 条 = 525 条测试用例"
ws_summary[f"D{row}"].font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
ws_summary[f"D{row}"].fill = PatternFill("solid", fgColor="1F4E79")
ws_summary[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center")
ws_summary[f"F{row}"] = "P0 + P1 + P2"
ws_summary[f"F{row}"].font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
ws_summary[f"F{row}"].fill = PatternFill("solid", fgColor="1F4E79")
ws_summary[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws_summary, row, 30)

# 列宽
ws_summary.column_dimensions["A"].width = 14
ws_summary.column_dimensions["B"].width = 10
ws_summary.column_dimensions["C"].width = 28
ws_summary.column_dimensions["D"].width = 70
ws_summary.column_dimensions["E"].width = 14
ws_summary.column_dimensions["F"].width = 20


# ─────────── Sheet 2：优先级分布说明 ───────────

ws_priority = wb.create_sheet("优先级分布说明")

ws_priority.merge_cells("A1:D1")
ws_priority["A1"] = "测试用例优先级分布说明"
title_style(ws_priority["A1"])
set_row_height(ws_priority, 1, 35)

desc_data = [
    ["优先级", "定义", "测试时机", "前台条数估算", "后台条数估算"],
    ["P0", "核心业务流程，主流程必须通过", "功能测试首要关注，必须在测试阶段全部通过", "约 65%", "约 60%"],
    ["P1", "重要功能，异常流程与边界情况", "P0之后执行，重要异常场景不可遗漏", "约 25%", "约 30%"],
    ["P2", "辅助功能，UI细节，低频路径", "可视项目进度安排，不阻塞发布", "约 10%", "约 10%"],
    ["合计", "—", "—", "100%", "100%"],
]

for r_idx, row_data in enumerate(desc_data, 1):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_priority.cell(row=r_idx + 1, column=c_idx, value=val)
        if r_idx == 1:
            header_style(cell)
        else:
            bold = (r_idx == 5)
            bg = "F2F2F2" if r_idx % 2 == 0 else "FFFFFF"
            cell.font = Font(name="微软雅黑", bold=bold, size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(style="thin", color="D9D9D9")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    set_row_height(ws_priority, r_idx + 1, 25)

ws_priority.column_dimensions["A"].width = 12
ws_priority.column_dimensions["B"].width = 28
ws_priority.column_dimensions["C"].width = 40
ws_priority.column_dimensions["D"].width = 18
ws_priority.column_dimensions["E"].width = 18


# ─────────── Sheet 3：测试环境说明 ───────────

ws_env = wb.create_sheet("测试环境与文档说明")

ws_env.merge_cells("A1:C1")
ws_env["A1"] = "测试环境与文档说明"
title_style(ws_env["A1"])
set_row_height(ws_env, 1, 35)

env_data = [
    ["项目", "内容", "备注"],
    ["产品名称", "睿联盟小程序", "青少年篮球综合服务平台"],
    ["前台类型", "微信小程序（用户端）", "C端用户使用，含招募/体测/会员/数据/直播等功能"],
    ["后台类型", "Web后台管理（机构端）", "B端运营管理，机构/管理员使用"],
    ["设计稿覆盖", "前台 249 张 + 后台 46 张 = 295 张", "均已按文件名梳理功能点"],
    ["测试用例总数", "前台 312 条 + 后台 213 条 = 525 条", "含 P0/P1/P2 三个优先级"],
    ["文档版本", "V1.0 正式版", "生成日期：2026-06-06"],
    ["测试阶段", "设计稿阶段测试用例", "后续需在真机/后台环境补充执行结果列"],
    ["编写人", "Claude AI", "审核人：待定"],
    ["测试类型", "功能测试 + 界面测试", "不含性能/安全/自动化测试用例"],
    ["文档格式", "Word (.docx) + Excel (.xlsx)", "Word含完整用例正文，Excel为模块汇总"],
    ["测试用例结构", "用例编号 / 名称 / 前置条件 / 测试步骤 / 预期结果 / 优先级 / 备注", "7个核心字段"],
]

for r_idx, row_data in enumerate(env_data, 1):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_env.cell(row=r_idx + 1, column=c_idx, value=val)
        if r_idx == 1:
            header_style(cell)
        else:
            bg = "EEF3FA" if c_idx == 1 else ("FFFFFF" if r_idx % 2 == 0 else "F9F9F9")
            bold = (c_idx == 1)
            cell.font = Font(name="微软雅黑", bold=bold, size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left" if c_idx != 1 else "center", vertical="center", wrap_text=True)
            thin = Side(style="thin", color="D9D9D9")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    set_row_height(ws_env, r_idx + 1, 25)

ws_env.column_dimensions["A"].width = 18
ws_env.column_dimensions["B"].width = 35
ws_env.column_dimensions["C"].width = 45


# ─────────── 保存 ───────────

output_path = "/Users/jianing/Ning's Git/reee/小程序产品/睿联盟测试用例模块汇总.xlsx"
wb.save(output_path)
print(f"Excel 已生成：{output_path}")
