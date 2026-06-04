"""
TikTok Shop Data Analytics
Main analysis script

数据处理流程:
1. 加载产品成本表 (根目录) → 提取seller_sku对应的成本信息
2. 加载订单数据 (日期文件夹) → 筛选有效订单(非取消、Normal)
3. 订单 + 产品成本表 → 通过Seller SKU关联
4. 计算利润 = 收入 - 产品成本 - 佣金 - 物流费
5. GMV广告数据 → 计算广告ROI
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
import openpyxl

# ============ 常量配置 ============
THB_TO_CNY = 4.7  # 泰铢兑人民币汇率
USD_TO_CNY = 6.8    # 美元兑人民币汇率

# 平台佣金费率
PLATFORM_COMMISSION_RATE = 0.032   # 平台佣金 3.2%
PLATFORM_GROWTH_RATE = 0.0809     # 平台增长费 8.09%
EXCHANGE_LOSS_RATE = 0.01         # 汇率损耗 1%
ORDER_FIXED_FEE_THB = 1.05       # 每笔订单固定处理费 (泰铢)

# ============ 路径配置 ============
PROJECT_ROOT = Path(__file__).parent.parent
DATA_DIR = PROJECT_ROOT / "data"
REPORTS_DIR = PROJECT_ROOT / "reports"

# ============ 样式定义 ============
FONT_NAME = '等线'
FONT_SIZE = 11
FONT_SIZE_TITLE = 14

# 颜色
HEADER_FILL_COLOR = 'D9E1F2'  # 浅蓝色

# 边框
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 数字格式
FMT_MONEY = '#,##0.0'
FMT_PERCENT = '0%'
FMT_QUANTITY = '#,##0.0'


def apply_header_style(cell):
    """应用表头样式"""
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    cell.fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type='solid')
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def display_header(text):
    """表头单位括号行内换行,例如 总收入(CNY) -> 总收入\n(CNY)。"""
    if isinstance(text, str) and '(' in text and ')' in text:
        return text.replace('(', '\n(')
    return text


def apply_data_style(cell, fmt=FMT_MONEY):
    """应用数据样式"""
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
    cell.border = THIN_BORDER
    cell.number_format = fmt


# ============ 数据加载 ============

def load_product_cost_table():
    """
    加载产品成本表
    data目录下的 产品成本表-XXXX.xlsx
    返回: 包含seller_sku、出厂价、海运物流单件成本、包装的DataFrame
    """
    # 查找data目录的成本表
    cost_files = list(DATA_DIR.glob("产品成本表-*.xlsx"))
    if not cost_files:
        raise FileNotFoundError("未找到产品成本表文件")

    # 使用最新的成本表
    latest_cost_file = sorted(cost_files)[-1]
    df = pd.read_excel(latest_cost_file, engine='openpyxl')

    # 提取关键列
    cost_df = df[['Seller SKU', '产品名称',
                  '出厂价',
                  '海运物流\n单件成本',
                  '包装']].copy()

    # 重命名列 - 统一为小写seller_sku便于匹配
    cost_df = cost_df.rename(columns={
        'Seller SKU': 'seller_sku',
        '产品名称': 'product_name_cn',
        '出厂价': 'factory_price',
        '海运物流\n单件成本': 'shipping_cost',
        '包装': 'package_cost'
    })

    # 转换数值类型
    for col in ['factory_price', 'shipping_cost', 'package_cost']:
        cost_df[col] = pd.to_numeric(cost_df[col], errors='coerce').fillna(0)

    # 清理: 删除重复的seller_sku和空值
    cost_df = cost_df.dropna(subset=['seller_sku'])
    cost_df = cost_df.drop_duplicates(subset=['seller_sku'], keep='first')

    return cost_df


def load_product_management():
    """
    加载产品管理表
    文件名: Tiktoksellercenter产品管理表*.xlsx
    返回: seller_sku → sku_id 映射
    """
    pm_files = list(DATA_DIR.glob("Tiktoksellercenter产品管理表*.xlsx"))
    if not pm_files:
        return pd.DataFrame()  # 返回空DF

    latest_pm_file = sorted(pm_files)[-1]
    # 跳过前两行(元数据行),使用第0行作为标题
    df = pd.read_excel(latest_pm_file, header=0, skiprows=[1, 2], engine='openpyxl')

    pm_df = df[['seller_sku', 'sku_id']].dropna(subset=['seller_sku'])
    # 去重
    pm_df = pm_df.drop_duplicates(subset=['seller_sku'], keep='first')
    return pm_df


def load_order_data(date_folder):
    """
    加载订单数据
    date_folder: 日期文件夹名,如 "20260519"
    返回: 所有订单数据(含已取消和normal、空值状态)
    """
    folder_path = DATA_DIR / date_folder
    order_files = list(folder_path.glob("*全部订单*.xlsx"))
    # 过滤掉临时文件(以~$开头)
    order_files = [f for f in order_files if not f.name.startswith('~$')]

    if not order_files:
        raise FileNotFoundError(f"未找到订单文件: {folder_path}")

    latest_order_file = sorted(order_files)[-1]
    # 使用openpyxl读取Excel,跳过前两行元数据,第3行作为标题
    wb = openpyxl.load_workbook(latest_order_file)
    ws = wb.active

    # 获取第3行开始的数据
    data = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        data.append(row)

    # 获取第1行作为标题
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    df = pd.DataFrame(data, columns=headers)

    return df, latest_order_file.name


def preprocess_orders(df):
    """
    预处理订单数据:
    1. 清理坏数据行
    2. 分离: Normal订单 vs 样品订单(空值)
    注意: 不做去重,每行代表一个订单项(同一Order ID可能有多行,代表不同商品)
    """
    # 清理坏数据行 (如 "Current order status.")
    df_clean = df[df['Order Status'] != 'Current order status.'].copy()

    # 分离Normal订单和样品订单(空值)
    df_normal = df_clean[df_clean['Normal or Pre-order'] == 'Normal'].copy()
    df_sample = df_clean[df_clean['Normal or Pre-order'].isna()].copy()

    return df_normal, df_sample


def load_gmv_data(date_folder):
    """
    加载GMV广告投放数据
    date_folder: 日期文件夹名
    返回: GMV数据DataFrame
    """
    folder_path = DATA_DIR / date_folder
    gmv_files = list(folder_path.glob("*gmv*.xlsx"))

    if not gmv_files:
        return pd.DataFrame()

    latest_gmv_file = sorted(gmv_files)[-1]
    df = pd.read_excel(latest_gmv_file, engine='openpyxl')

    return df, latest_gmv_file.name


# ============ 数据处理 ============

def enrich_orders_with_cost(order_df, cost_df, pm_df):
    """
    将订单数据与成本表关联
    通过 Seller SKU 匹配,添加:
    - 产品名称(中文)
    - 出厂价
    - 海运物流单件成本
    - 包装成本
    """
    # 统一列名大小写 - 订单表用 "Seller SKU",成本表用 "seller_sku"
    # 转换为统一格式便于匹配
    order_df = order_df.rename(columns={'Seller SKU': 'seller_sku'})

    # 合并产品成本信息
    df = order_df.merge(cost_df, on='seller_sku', how='left')

    # 合并产品管理表信息
    if not pm_df.empty:
        # 重命名pm_df的列以避免冲突
        pm_df = pm_df.rename(columns={'sku_id': 'sku_id_from_pm'})
        df = df.merge(pm_df, on='seller_sku', how='left')

    return df


def calculate_normal_order_profit(df):
    """
    计算Normal订单的利润

    收入 = (买家实付总额 + 平台折扣) (泰铢) / 4.7 → 人民币
    其中 平台折扣 = SKU Platform Discount + Shipping Fee Platform Discount

    成本项目:
    - 产品成本 = (出厂价 + 海运物流单件成本) × Quantity
    - 包装成本 = 包装费 (单件,不需要×Quantity)
    - 平台佣金 = (买家实付总额 + 平台折扣)×(3.2%+8.09%+1%) + 1.05泰铢
    - 包邮服务 = (买家实付总额 + 平台折扣)×7.49%

    订单成本 = 产品成本 + 包装成本 + 平台佣金 + 包邮服务

    利润 = 收入 - 订单成本
    """
    # 转换数值类型
    df['Order Amount'] = pd.to_numeric(df['Order Amount'], errors='coerce').fillna(0)

    # 读取平台折扣列
    df['sku_platform_discount'] = pd.to_numeric(df.get('SKU Platform Discount'), errors='coerce').fillna(0)
    df['shipping_fee_platform_discount'] = pd.to_numeric(df.get('Shipping Fee Platform Discount'), errors='coerce').fillna(0)

    # 最终订单金额 = 原始订单金额 + 平台折扣
    df['final_order_amount_thb'] = df['Order Amount'] + df['sku_platform_discount'] + df['shipping_fee_platform_discount']

    # 收入 (泰铢 → 人民币), 使用最终订单金额
    df['income_cny'] = df['final_order_amount_thb'] / THB_TO_CNY

    # 确保数值类型
    for col in ['factory_price', 'shipping_cost', 'package_cost']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)

    # 产品成本 = (出厂价 + 海运物流单件成本) × Quantity
    df['product_cost_cny'] = (
        (df['factory_price'] + df['shipping_cost']) * df['Quantity']
    )

    # 包装成本 (单件,不需要×Quantity)
    df['package_cost_cny'] = df['package_cost']

    # 平台佣金 = (买家实付总额 + 平台折扣)×(3.2%+8.09%+1%) + 1.05泰铢
    commission_rate = PLATFORM_COMMISSION_RATE + PLATFORM_GROWTH_RATE + EXCHANGE_LOSS_RATE  # 12.29%
    df['commission_cny'] = (
        df['final_order_amount_thb'] * commission_rate + ORDER_FIXED_FEE_THB
    ) / THB_TO_CNY

    # 包邮服务 = (买家实付总额 + 平台折扣)×7.49%
    FREE_SHIPPING_RATE = 0.0749
    df['free_shipping_cny'] = (df['final_order_amount_thb'] * FREE_SHIPPING_RATE) / THB_TO_CNY

    # 订单成本 = 产品成本 + 包装成本 + 平台佣金 + 包邮服务
    df['order_cost_cny'] = (
        df['product_cost_cny'] +
        df['package_cost_cny'] +
        df['commission_cny'] +
        df['free_shipping_cny']
    )

    # 利润 = 收入 - 订单成本
    df['profit_cny'] = df['income_cny'] - df['order_cost_cny']

    # 利润(泰铢) - 用于导出
    df['profit_thb'] = df['profit_cny'] * THB_TO_CNY

    # 利润率 (数值比例,Excel中按百分比显示)
    df['profit_rate'] = df.apply(
        lambda x: x['profit_cny'] / x['income_cny'] if x['income_cny'] > 0 else 0,
        axis=1
    )

    # 单件利润
    df['unit_profit_cny'] = df.apply(
        lambda x: x['profit_cny'] / x['Quantity'] if x['Quantity'] > 0 else 0,
        axis=1
    )

    # 订单金额泰铢 (原始数据)
    df['order_amount_thb'] = df['Order Amount']

    return df


def calculate_sample_order_profit(df):
    """
    计算样品订单的利润

    样品订单没有收入,只有成本:
    - 产品成本 = (出厂价 + 海运物流单件成本) × Quantity
    - 包装成本 = 包装费 (单件)
    - 固定快递费 = 7元

    订单成本 = 产品成本 + 包装成本 + 固定快递费

    利润 = 0 - 订单成本 = -订单成本
    """
    # 确保数值类型
    for col in ['factory_price', 'shipping_cost', 'package_cost']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)

    # 样品没有收入
    df['income_cny'] = 0.0
    df['order_amount_thb'] = 0.0
    df['final_order_amount_thb'] = 0.0
    df['sku_platform_discount'] = 0.0
    df['shipping_fee_platform_discount'] = 0.0

    # 产品成本 = (出厂价 + 海运物流单件成本) × Quantity
    df['product_cost_cny'] = (
        (df['factory_price'] + df['shipping_cost']) * df['Quantity']
    )

    # 包装成本 (单件)
    df['package_cost_cny'] = df['package_cost']

    # 固定快递费 7元
    FIXED_SHIPPING_CNY = 7.0

    # 样品订单成本 = 产品成本 + 包装成本 + 固定快递费
    df['order_cost_cny'] = (
        df['product_cost_cny'] +
        df['package_cost_cny'] +
        FIXED_SHIPPING_CNY
    )

    # 样品没有平台佣金和包邮服务
    df['commission_cny'] = 0.0
    df['free_shipping_cny'] = 0.0

    # 利润 = 0 - 订单成本
    df['profit_cny'] = -df['order_cost_cny']

    # 利润(泰铢) - 用于导出 (样品利润为负)
    df['profit_thb'] = df['profit_cny'] * THB_TO_CNY

    # 利润率 (样品订单利润为负,标记为-N/A或特殊标记)
    df['profit_rate'] = '-'

    # 单件利润
    df['unit_profit_cny'] = df.apply(
        lambda x: x['profit_cny'] / x['Quantity'] if x['Quantity'] > 0 else 0,
        axis=1
    )

    return df


def calculate_gmv_roi(gmv_df):
    """
    计算GMV广告ROI
    成本和收入都是美元,转人民币计算
    """
    if gmv_df.empty:
        return gmv_df

    # 成本和收入 (美元 → 人民币)
    gmv_df['cost_cny'] = gmv_df['成本'] * USD_TO_CNY
    gmv_df['revenue_cny'] = gmv_df['总收入（当前店铺）'] * USD_TO_CNY

    # ROI = 收入 / 成本
    gmv_df['roi'] = gmv_df.apply(
        lambda x: x['revenue_cny'] / x['cost_cny'] if x['cost_cny'] > 0 else 0,
        axis=1
    )

    # 每单成本
    gmv_df['cost_per_order_cny'] = gmv_df.apply(
        lambda x: x['cost_cny'] / x['SKU 订单数（当前店铺）'] if x['SKU 订单数（当前店铺）'] > 0 else 0,
        axis=1
    )

    return gmv_df


# ============ 分析报告 ============

def generate_order_report(df, date_folder):
    """
    生成订单分析报告
    """
    report = {
        'report_date': date_folder,
        'total_orders': len(df),
        'total_quantity': df['Quantity'].sum(),
        'total_income_cny': df['income_cny'].sum(),
        'total_product_cost_cny': df['product_cost_cny'].sum(),
        'total_commission_cny': df['commission_cny'].sum(),
        'total_order_cost_cny': df['order_cost_cny'].sum(),
        'total_profit_cny': df['profit_cny'].sum(),
        'avg_profit_per_order': df['profit_cny'].mean(),
        'profit_rate': df['profit_cny'].sum() / df['income_cny'].sum() if df['income_cny'].sum() > 0 else 0,
    }

    # 产品利润排行
    product_profit = df.groupby('seller_sku').agg({
        'Quantity': 'sum',
        'income_cny': 'sum',
        'product_cost_cny': 'sum',
        'commission_cny': 'sum',
        'order_cost_cny': 'sum',
        'profit_cny': 'sum',
        'Product Name': 'first',
        'factory_price': 'first',
    }).reset_index()
    product_profit = product_profit.sort_values('profit_cny', ascending=False)

    return report, product_profit


def generate_gmv_report(df):
    """
    生成GMV广告分析报告
    """
    if df.empty:
        return {'total_cost_cny': 0, 'total_revenue_cny': 0, 'total_orders': 0, 'avg_roi': 0}

    report = {
        'total_cost_cny': df['cost_cny'].sum(),
        'total_revenue_cny': df['revenue_cny'].sum(),
        'total_orders': df['SKU 订单数（当前店铺）'].sum(),
        'avg_roi': df['roi'].mean(),
        'avg_cost_per_order': df['cost_per_order_cny'].mean(),
    }

    # 广告计划效果排行 - 如果有'广告计划名称'列则分组,否则按天汇总
    if '广告计划名称' in df.columns:
        campaign_performance = df.groupby('广告计划名称').agg({
            'cost_cny': 'sum',
            'revenue_cny': 'sum',
            'SKU 订单数（当前店铺）': 'sum',
            'roi': 'mean',
        }).reset_index()
        campaign_performance = campaign_performance.sort_values('roi', ascending=False)
    else:
        # 按天汇总
        campaign_performance = df.groupby('按天').agg({
            'cost_cny': 'sum',
            'revenue_cny': 'sum',
            'SKU 订单数（当前店铺）': 'sum',
            'roi': 'mean',
        }).reset_index()
        campaign_performance = campaign_performance.sort_values('按天', ascending=False)

    return report, campaign_performance


def export_order_master_table(df_normal, df_sample, date_folder, gmv_revenue_cny=0, gmv_cost_cny=0, gmv_roi=0):
    """
    导出订单收入大表到 data/{date}/ 目录

    包含5个Sheet:
    1. 总览 - 所有订单汇总(含GMV数据)
    2. 产品统计(Normal) - Normal订单按产品汇总
    3. 产品统计(样品) - 样品订单按产品汇总
    4. Normal订单 - 正常订单明细
    5. 样品订单 - 达人样品订单明细

    - OpenPyXL styling: thin borders on all cells, header fill with light blue (D9E1F2)
    - All sheets have Excel formulas where calculations happen
    - Number formatting: money uses '#,##0.00', percentages use '0%', quantities use '#,##0.0'
    """
    output_dir = DATA_DIR / date_folder
    output_dir.mkdir(exist_ok=True, parents=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"订单收入大表_{date_folder}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # ========== Sheet 1: 总览 ==========
    ws_summary = wb.create_sheet('总览')

    # 合并标题单元格
    ws_summary.merge_cells('A1:I1')
    title_cell = ws_summary.cell(row=1, column=1, value='TikTok订单数据汇总')
    title_cell.font = Font(name=FONT_NAME, size=FONT_SIZE_TITLE, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 表头行 - Row 3: 中文, Row 4: 英文
    chn_headers = ['类别', '订单数', '总收入(CNY)', '总收入(泰铢)',
                   '总成本(CNY)', '总成本(泰铢)', '总利润(CNY)', '总利润(泰铢)', '利润率(%)']
    eng_headers = ['Type', 'Order Count', 'Total Revenue(CNY)', 'Total Revenue(THB)',
                   'Total Cost(CNY)', 'Total Cost(THB)', 'Total Profit(CNY)', 'Total Profit(THB)', 'Profit Rate(%)']

    for col_idx, (chn, eng) in enumerate(zip(chn_headers, eng_headers), start=1):
        chn_cell = ws_summary.cell(row=3, column=col_idx, value=display_header(chn))
        eng_cell = ws_summary.cell(row=4, column=col_idx, value=display_header(eng))
        apply_header_style(chn_cell)
        apply_header_style(eng_cell)

    # 数据行 - Row 5: Normal订单, Row 6: 样品订单, Row 7: 全部累加, Row 8: GMV广告, Row 9: 项目利润率
    # 计算汇总数据
    total_income = df_normal['income_cny'].sum() + df_sample['income_cny'].sum()
    total_cost = df_normal['order_cost_cny'].sum() + df_sample['order_cost_cny'].sum()
    total_profit = df_normal['profit_cny'].sum() + df_sample['profit_cny'].sum()

    normal_count = len(df_normal)
    normal_income = df_normal['income_cny'].sum()
    normal_cost = df_normal['order_cost_cny'].sum()
    normal_profit = df_normal['profit_cny'].sum()

    sample_count = len(df_sample)
    sample_cost = df_sample['order_cost_cny'].sum()
    sample_profit = df_sample['profit_cny'].sum()

    # Row 5: Normal订单 - 使用公式
    ws_summary.cell(row=5, column=1, value='Normal订单')
    ws_summary.cell(row=5, column=2, value=normal_count)
    ws_summary.cell(row=5, column=3, value=normal_income)  # 总收入CNY
    ws_summary.cell(row=5, column=4, value='=C5*4.7')  # 总收入THB = CNY * 4.7
    ws_summary.cell(row=5, column=5, value=normal_cost)  # 总成本CNY
    ws_summary.cell(row=5, column=6, value='=E5*4.7')  # 总成本THB
    ws_summary.cell(row=5, column=7, value=normal_profit)  # 总利润CNY
    ws_summary.cell(row=5, column=8, value='=G5*4.7')  # 总利润THB
    ws_summary.cell(row=5, column=9, value='=IF(C5>0,G5/C5,0)')  # 利润率

    # Row 6: 样品订单
    ws_summary.cell(row=6, column=1, value='样品订单')
    ws_summary.cell(row=6, column=2, value=sample_count)
    ws_summary.cell(row=6, column=3, value=0)  # 无收入
    ws_summary.cell(row=6, column=4, value=0)
    ws_summary.cell(row=6, column=5, value=sample_cost)
    ws_summary.cell(row=6, column=6, value='=E6*4.7')
    ws_summary.cell(row=6, column=7, value=sample_profit)
    ws_summary.cell(row=6, column=8, value='=G6*4.7')
    ws_summary.cell(row=6, column=9, value='-')  # 无利润率

    # Row 7: 全部累加
    ws_summary.cell(row=7, column=1, value='全部累加')
    ws_summary.cell(row=7, column=2, value='=B5+B6')
    ws_summary.cell(row=7, column=3, value='=C5+C6')
    ws_summary.cell(row=7, column=4, value='=D5+D6')
    ws_summary.cell(row=7, column=5, value='=E5+E6')
    ws_summary.cell(row=7, column=6, value='=F5+F6')
    ws_summary.cell(row=7, column=7, value='=G5+G6')
    ws_summary.cell(row=7, column=8, value='=H5+H6')
    ws_summary.cell(row=7, column=9, value='=IF(C7>0,G7/C7,0)')

    # Row 8: GMV广告 - GMV收入单独列示(不计入项目总收入),GMV成本计入项目总成本
    ws_summary.cell(row=8, column=1, value='GMV广告')
    ws_summary.cell(row=8, column=2, value='-')
    ws_summary.cell(row=8, column=3, value=gmv_revenue_cny)  # GMV收入(单独列示)
    ws_summary.cell(row=8, column=4, value='=C8*4.7')  # GMV收入THB
    ws_summary.cell(row=8, column=5, value=gmv_cost_cny)  # GMV成本
    ws_summary.cell(row=8, column=6, value='=E8*4.7')  # GMV成本THB
    ws_summary.cell(row=8, column=7, value='=C8-E8')  # GMV利润
    ws_summary.cell(row=8, column=8, value='=G8*4.7')  # GMV利润THB
    ws_summary.cell(row=8, column=9, value=gmv_roi if gmv_roi > 0 else 0)

    # Row 9: 项目利润率
    # 项目总收入 = 订单总收入（不含GMV收入）
    # 项目总成本 = 订单总成本 + GMV成本
    # 项目总利润 = 项目总收入 - 项目总成本
    ws_summary.cell(row=9, column=1, value='项目利润率')
    ws_summary.cell(row=9, column=2, value='-')
    ws_summary.cell(row=9, column=3, value='=C7')  # 项目总收入 = 订单总收入,不含GMV收入
    ws_summary.cell(row=9, column=4, value='=C9*4.7')
    ws_summary.cell(row=9, column=5, value='=E7+E8')  # 项目总成本 = 订单成本 + GMV成本
    ws_summary.cell(row=9, column=6, value='=E9*4.7')
    ws_summary.cell(row=9, column=7, value='=C9-E9')  # 项目总利润 = 订单收入 - 订单成本 - GMV成本
    ws_summary.cell(row=9, column=8, value='=G9*4.7')
    ws_summary.cell(row=9, column=9, value='=IF(E9>0,G9/E9,0)')  # 项目利润率

    # 应用数据样式 - CNY列(总收入/总成本/总利润)和利润率加粗
    for row in range(5, 10):
        for col in range(1, 10):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [3, 5, 7, 9]:  # CNY列和利润率加粗
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            else:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
            if col in [3, 5, 7]:  # CNY columns - 1位小数
                cell.number_format = '#,##0.0'
            elif col in [4, 6, 8]:  # THB columns - 1位小数
                cell.number_format = '#,##0.0'
            elif col == 9:  # 利润率 - 取整百分比
                cell.number_format = '0%'
            elif col == 2:  # 订单数
                cell.number_format = '#,##0'

    # 调整列宽
    for col in range(1, 10):
        ws_summary.column_dimensions[get_column_letter(col)].width = 14
    ws_summary.row_dimensions[3].height = 36
    ws_summary.row_dimensions[4].height = 36

    # ========== Sheet 2: 产品统计(Normal) ==========
    ws_product_normal = wb.create_sheet('产品统计(Normal)')

    # 聚合Normal订单数据
    product_stats_normal = df_normal.groupby('seller_sku').agg({
        'product_name_cn': 'first',
        'Quantity': 'sum',
        'income_cny': 'sum',
        'product_cost_cny': 'sum',
        'package_cost_cny': 'sum',
        'commission_cny': 'sum',
        'free_shipping_cny': 'sum',
        'order_cost_cny': 'sum',
        'profit_cny': 'sum',
    }).reset_index()

    product_stats_normal = product_stats_normal.sort_values('Quantity', ascending=False)

    # 表头
    prod_normal_headers = ['商家SKU', '产品名称(中文)', '订单数',
                          '总收入(CNY)', '总收入(泰铢)',
                          '产品成本(CNY)', '包装成本(CNY)', '平台佣金(CNY)', '包邮服务(CNY)',
                          '订单总成本(CNY)', '订单总成本(泰铢)',
                          '总利润(CNY)', '总利润(泰铢)', '单品利润率']
    for col_idx, header in enumerate(prod_normal_headers, start=1):
        cell = ws_product_normal.cell(row=1, column=col_idx, value=display_header(header))
        apply_header_style(cell)

    # 数据行 - 使用公式
    for row_idx, (_, row_data) in enumerate(product_stats_normal.iterrows(), start=2):
        ws_product_normal.cell(row=row_idx, column=1, value=row_data['seller_sku'])
        ws_product_normal.cell(row=row_idx, column=2, value=row_data['product_name_cn'])
        ws_product_normal.cell(row=row_idx, column=3, value=row_data['Quantity'])
        ws_product_normal.cell(row=row_idx, column=4, value=row_data['income_cny'])  # 收入CNY
        ws_product_normal.cell(row=row_idx, column=5, value=f'=D{row_idx}*4.7')  # 收入THB = CNY * 4.7
        ws_product_normal.cell(row=row_idx, column=6, value=row_data['product_cost_cny'])
        ws_product_normal.cell(row=row_idx, column=7, value=row_data['package_cost_cny'])
        ws_product_normal.cell(row=row_idx, column=8, value=row_data['commission_cny'])
        ws_product_normal.cell(row=row_idx, column=9, value=row_data['free_shipping_cny'])
        ws_product_normal.cell(row=row_idx, column=10, value=row_data['order_cost_cny'])
        ws_product_normal.cell(row=row_idx, column=11, value=f'=J{row_idx}*4.7')  # 成本THB = CNY * 4.7
        ws_product_normal.cell(row=row_idx, column=12, value=row_data['profit_cny'])
        ws_product_normal.cell(row=row_idx, column=13, value=f'=L{row_idx}*4.7')  # 利润THB = CNY * 4.7
        ws_product_normal.cell(row=row_idx, column=14, value=f'=IF(D{row_idx}>0,L{row_idx}/D{row_idx},0)')

    # 汇总行
    last_data_row = len(product_stats_normal) + 1
    summary_row = last_data_row + 1
    ws_product_normal.cell(row=summary_row, column=1, value='汇总')
    ws_product_normal.cell(row=summary_row, column=2, value='')
    ws_product_normal.cell(row=summary_row, column=3, value=f'=SUM(C2:C{last_data_row})')
    ws_product_normal.cell(row=summary_row, column=4, value=f'=SUM(D2:D{last_data_row})')
    ws_product_normal.cell(row=summary_row, column=5, value=f'=D{summary_row}*4.7')  # 汇总THB
    ws_product_normal.cell(row=summary_row, column=6, value=f'=SUM(F2:F{last_data_row})')
    ws_product_normal.cell(row=summary_row, column=7, value=f'=SUM(G2:G{last_data_row})')
    ws_product_normal.cell(row=summary_row, column=8, value=f'=SUM(H2:H{last_data_row})')
    ws_product_normal.cell(row=summary_row, column=9, value=f'=SUM(I2:I{last_data_row})')
    ws_product_normal.cell(row=summary_row, column=10, value=f'=SUM(J2:J{last_data_row})')
    ws_product_normal.cell(row=summary_row, column=11, value=f'=J{summary_row}*4.7')  # 汇总THB
    ws_product_normal.cell(row=summary_row, column=12, value=f'=SUM(L2:L{last_data_row})')
    ws_product_normal.cell(row=summary_row, column=13, value=f'=L{summary_row}*4.7')  # 汇总THB
    ws_product_normal.cell(row=summary_row, column=14, value=f'=IF(D{summary_row}>0,L{summary_row}/D{summary_row},0)')  # 利润率公式

    # 应用样式(包含汇总行) - CNY列(总收入/总成本/总利润)和利润率加粗
    for row in range(2, summary_row + 1):
        for col in range(1, 15):
            cell = ws_product_normal.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if row == summary_row:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            elif col == 1:  # SKU列左对齐
                cell.alignment = Alignment(horizontal='left')
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
            elif col in [4, 10, 12, 14]:  # CNY列和利润率加粗(总收入/总成本/总利润/利润率)
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            else:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
            if col in [4, 6, 7, 8, 9, 10, 12]:  # CNY columns - 1位小数
                cell.number_format = '#,##0.0'
            elif col in [5, 11, 13]:  # THB formula columns - 1位小数
                cell.number_format = '#,##0.0'
            elif col == 14:  # 利润率 - 取整百分比
                cell.number_format = '0%'
            elif col == 3:  # 订单数 - 1位小数
                cell.number_format = '#,##0.0'

    # 调整列宽 - 统一10,产品名称列18
    for col in range(1, 15):
        ws_product_normal.column_dimensions[get_column_letter(col)].width = 10
    ws_product_normal.column_dimensions['B'].width = 18
    ws_product_normal.row_dimensions[1].height = 36

    # ========== Sheet 3: 产品统计(样品) ==========
    ws_product_sample = wb.create_sheet('产品统计(样品)')

    # 聚合样品订单数据
    FIXED_SHIPPING_CNY = 7.0

    product_stats_sample = df_sample.groupby('seller_sku').agg({
        'product_name_cn': 'first',
        'Quantity': 'sum',
        'product_cost_cny': 'sum',
        'package_cost_cny': 'sum',
    }).reset_index()

    # 计算快递费 (每笔订单7元)
    sample_order_counts = df_sample.groupby('seller_sku').size()
    product_stats_sample['shipping_fee_cny'] = product_stats_sample['seller_sku'].map(sample_order_counts) * FIXED_SHIPPING_CNY

    # 重新计算订单总成本
    product_stats_sample['order_cost_cny'] = (
        product_stats_sample['product_cost_cny'] +
        product_stats_sample['package_cost_cny'] +
        product_stats_sample['shipping_fee_cny']
    )

    # 重新计算利润 (无收入,利润 = -成本)
    product_stats_sample['profit_cny'] = -product_stats_sample['order_cost_cny']

    product_stats_sample = product_stats_sample.sort_values('Quantity', ascending=False)

    # 表头
    prod_sample_headers = ['商家SKU', '产品名称(中文)', '订单数',
                          '产品成本(CNY)', '包装成本(CNY)', '快递费(CNY)',
                          '订单总成本(CNY)', '订单总成本(泰铢)',
                          '总利润(CNY)', '总利润(泰铢)']
    for col_idx, header in enumerate(prod_sample_headers, start=1):
        cell = ws_product_sample.cell(row=1, column=col_idx, value=display_header(header))
        apply_header_style(cell)

    # 数据行 - 使用公式
    for row_idx, (_, row_data) in enumerate(product_stats_sample.iterrows(), start=2):
        ws_product_sample.cell(row=row_idx, column=1, value=row_data['seller_sku'])
        ws_product_sample.cell(row=row_idx, column=2, value=row_data['product_name_cn'])
        ws_product_sample.cell(row=row_idx, column=3, value=row_data['Quantity'])
        ws_product_sample.cell(row=row_idx, column=4, value=row_data['product_cost_cny'])
        ws_product_sample.cell(row=row_idx, column=5, value=row_data['package_cost_cny'])
        ws_product_sample.cell(row=row_idx, column=6, value=row_data['shipping_fee_cny'])
        # 总成本 = D+E+F
        ws_product_sample.cell(row=row_idx, column=7, value=f'=D{row_idx}+E{row_idx}+F{row_idx}')
        ws_product_sample.cell(row=row_idx, column=8, value=f'=G{row_idx}*4.7')  # 成本THB
        # 利润 = -G (无收入)
        ws_product_sample.cell(row=row_idx, column=9, value=f'=-G{row_idx}')
        ws_product_sample.cell(row=row_idx, column=10, value=f'=I{row_idx}*4.7')  # 利润THB

    # 汇总行
    last_data_row_sample = len(product_stats_sample) + 1
    summary_row_sample = last_data_row_sample + 1
    ws_product_sample.cell(row=summary_row_sample, column=1, value='汇总')
    ws_product_sample.cell(row=summary_row_sample, column=2, value='')
    ws_product_sample.cell(row=summary_row_sample, column=3, value=f'=SUM(C2:C{last_data_row_sample})')
    ws_product_sample.cell(row=summary_row_sample, column=4, value=f'=SUM(D2:D{last_data_row_sample})')
    ws_product_sample.cell(row=summary_row_sample, column=5, value=f'=SUM(E2:E{last_data_row_sample})')
    ws_product_sample.cell(row=summary_row_sample, column=6, value=f'=SUM(F2:F{last_data_row_sample})')
    ws_product_sample.cell(row=summary_row_sample, column=7, value=f'=SUM(G2:G{last_data_row_sample})')
    ws_product_sample.cell(row=summary_row_sample, column=8, value=f'=G{summary_row_sample}*4.7')  # 汇总THB
    ws_product_sample.cell(row=summary_row_sample, column=9, value=f'=SUM(I2:I{last_data_row_sample})')
    ws_product_sample.cell(row=summary_row_sample, column=10, value=f'=I{summary_row_sample}*4.7')  # 汇总THB

    # 应用样式(包含汇总行) - CNY列(总成本/总利润)和利润率加粗
    for row in range(2, summary_row_sample + 1):
        for col in range(1, 11):
            cell = ws_product_sample.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if row == summary_row_sample:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            elif col == 1:  # SKU列左对齐
                cell.alignment = Alignment(horizontal='left')
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
            elif col in [7, 9, 10]:  # 总成本/总利润/利润率加粗
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            else:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
            if col in [4, 5, 6, 7, 9]:  # CNY columns - 1位小数
                cell.number_format = '#,##0.0'
            elif col in [8, 10]:  # THB formula columns - 1位小数
                cell.number_format = '#,##0.0'
            elif col == 3:  # 订单数 - 1位小数
                cell.number_format = '#,##0.0'

    # 调整列宽 - 统一10,产品名称列18
    for col in range(1, 11):
        ws_product_sample.column_dimensions[get_column_letter(col)].width = 10
    ws_product_sample.column_dimensions['B'].width = 18
    ws_product_sample.row_dimensions[1].height = 36

    # ========== Sheet 4: Normal订单 ==========
    ws_normal = wb.create_sheet('Normal订单')

    # 定义列
    normal_export_columns = [
        ('Order ID', '订单ID'),
        ('Order Status', '订单状态'),
        ('seller_sku', '商家SKU'),
        ('product_name_cn', '产品名称(中文)'),
        ('Product Name', '产品名称(泰语)'),
        ('Variation', '规格变体'),
        ('Quantity', '数量(件)'),
        ('order_amount_thb', '订单金额(泰铢)'),
        ('sku_platform_discount', 'SKU平台折扣(泰铢)'),
        ('shipping_fee_platform_discount', '运费平台折扣(泰铢)'),
        ('final_order_amount_thb', '最终订单金额(泰铢)'),
        ('income_cny', '收入(CNY)'),
        ('product_cost_cny', '产品成本(CNY)'),
        ('package_cost_cny', '包装成本(CNY)'),
        ('commission_cny', '平台佣金(CNY)'),
        ('free_shipping_cny', '包邮服务(CNY)'),
        ('order_cost_cny', '订单总成本(CNY)'),
        ('profit_cny', '利润(CNY)'),
        ('profit_thb', '利润(泰铢)'),
        ('profit_rate', '利润率(%)'),
        ('Created Time', '下单时间'),
    ]

    # 表头
    for col_idx, (col_key, col_name) in enumerate(normal_export_columns, start=1):
        cell = ws_normal.cell(row=1, column=col_idx, value=display_header(col_name))
        apply_header_style(cell)

    # 准备数据
    available_cols = [c[0] for c in normal_export_columns if c[0] in df_normal.columns]
    df_normal_export = df_normal[available_cols].copy()

    # 数据行
    for row_idx, (_, row_data) in enumerate(df_normal_export.iterrows(), start=2):
        for col_idx, (col_key, _) in enumerate(normal_export_columns, start=1):
            if col_key in df_normal.columns:
                cell = ws_normal.cell(row=row_idx, column=col_idx, value=row_data[col_key])
            else:
                cell = ws_normal.cell(row=row_idx, column=col_idx, value='')

    # 应用样式 - 收入/总成本/利润/利润率加粗
    money_cols = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]
    last_normal_data_row = len(df_normal_export) + 1
    summary_normal_row = last_normal_data_row + 1

    for row in range(2, summary_normal_row + 1):
        for col in range(1, len(normal_export_columns) + 1):
            cell = ws_normal.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if row == summary_normal_row:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            elif col in [12, 17, 18, 20]:  # 收入/总成本/利润/利润率加粗
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            else:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
            if col in money_cols:
                cell.number_format = '#,##0.0'
            elif col == 20:
                cell.number_format = '0%'
            elif col == 7:  # Quantity
                cell.number_format = '#,##0.0'

    # 汇总行 - Normal订单
    ws_normal.cell(row=summary_normal_row, column=1, value='汇总')
    ws_normal.cell(row=summary_normal_row, column=2, value='')
    ws_normal.cell(row=summary_normal_row, column=3, value=f'=SUM(G2:G{last_normal_data_row})')  # 订单数
    ws_normal.cell(row=summary_normal_row, column=4, value='')
    ws_normal.cell(row=summary_normal_row, column=5, value='')
    ws_normal.cell(row=summary_normal_row, column=6, value='')
    ws_normal.cell(row=summary_normal_row, column=7, value=f'=SUM(G2:G{last_normal_data_row})')  # 数量
    ws_normal.cell(row=summary_normal_row, column=8, value=f'=SUM(H2:H{last_normal_data_row})')  # 订单金额
    ws_normal.cell(row=summary_normal_row, column=9, value=f'=SUM(I2:I{last_normal_data_row})')  # SKU平台折扣
    ws_normal.cell(row=summary_normal_row, column=10, value=f'=SUM(J2:J{last_normal_data_row})')  # 运费平台折扣
    ws_normal.cell(row=summary_normal_row, column=11, value=f'=SUM(K2:K{last_normal_data_row})')  # 最终订单金额
    ws_normal.cell(row=summary_normal_row, column=12, value=f'=SUM(L2:L{last_normal_data_row})')  # 收入
    ws_normal.cell(row=summary_normal_row, column=13, value=f'=SUM(M2:M{last_normal_data_row})')  # 产品成本
    ws_normal.cell(row=summary_normal_row, column=14, value=f'=SUM(N2:N{last_normal_data_row})')  # 包装成本
    ws_normal.cell(row=summary_normal_row, column=15, value=f'=SUM(O2:O{last_normal_data_row})')  # 平台佣金
    ws_normal.cell(row=summary_normal_row, column=16, value=f'=SUM(P2:P{last_normal_data_row})')  # 包邮服务
    ws_normal.cell(row=summary_normal_row, column=17, value=f'=SUM(Q2:Q{last_normal_data_row})')  # 订单总成本
    ws_normal.cell(row=summary_normal_row, column=18, value=f'=SUM(R2:R{last_normal_data_row})')  # 利润
    ws_normal.cell(row=summary_normal_row, column=19, value=f'=SUM(S2:S{last_normal_data_row})')  # 利润泰铢
    ws_normal.cell(row=summary_normal_row, column=20, value=f'=IF(L{summary_normal_row}>0,R{summary_normal_row}/L{summary_normal_row},0)')  # 利润率

    # 调整列宽 - 统一10,产品名称列18
    for col in range(1, len(normal_export_columns) + 1):
        ws_normal.column_dimensions[get_column_letter(col)].width = 10
    ws_normal.column_dimensions['D'].width = 18
    ws_normal.column_dimensions['E'].width = 18
    ws_normal.row_dimensions[1].height = 36

    # ========== Sheet 5: 样品订单 ==========
    ws_sample = wb.create_sheet('样品订单')

    # 定义列 - 样品订单没有收入相关列
    sample_export_columns = [
        ('Order ID', '订单ID'),
        ('Order Status', '订单状态'),
        ('seller_sku', '商家SKU'),
        ('product_name_cn', '产品名称(中文)'),
        ('Product Name', '产品名称(泰语)'),
        ('Variation', '规格变体'),
        ('Quantity', '数量(件)'),
        ('product_cost_cny', '产品成本(CNY)'),
        ('package_cost_cny', '包装成本(CNY)'),
        ('order_cost_cny', '订单总成本(CNY)'),
        ('profit_cny', '利润(CNY)'),
        ('profit_thb', '利润(泰铢)'),
        ('Created Time', '下单时间'),
    ]

    # 表头
    for col_idx, (col_key, col_name) in enumerate(sample_export_columns, start=1):
        cell = ws_sample.cell(row=1, column=col_idx, value=display_header(col_name))
        apply_header_style(cell)

    # 准备数据
    sample_available_cols = [c[0] for c in sample_export_columns if c[0] in df_sample.columns]
    df_sample_export = df_sample[sample_available_cols].copy()

    # 数据行
    for row_idx, (_, row_data) in enumerate(df_sample_export.iterrows(), start=2):
        for col_idx, (col_key, _) in enumerate(sample_export_columns, start=1):
            if col_key in df_sample.columns:
                cell = ws_sample.cell(row=row_idx, column=col_idx, value=row_data[col_key])
            else:
                cell = ws_sample.cell(row=row_idx, column=col_idx, value='')

    # 应用样式 - 总成本/利润/利润率加粗
    sample_money_cols = [8, 9, 10, 11, 12]
    last_sample_data_row = len(df_sample_export) + 1
    summary_sample_row = last_sample_data_row + 1

    for row in range(2, summary_sample_row + 1):
        for col in range(1, len(sample_export_columns) + 1):
            cell = ws_sample.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if row == summary_sample_row:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            elif col in [10, 11, 12]:  # 总成本/利润加粗
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            else:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
            if col in sample_money_cols:
                cell.number_format = '#,##0.0'
            elif col == 7:  # Quantity
                cell.number_format = '#,##0.0'

    # 汇总行 - 样品订单
    ws_sample.cell(row=summary_sample_row, column=1, value='汇总')
    ws_sample.cell(row=summary_sample_row, column=2, value='')
    ws_sample.cell(row=summary_sample_row, column=3, value='')
    ws_sample.cell(row=summary_sample_row, column=4, value='')
    ws_sample.cell(row=summary_sample_row, column=5, value='')
    ws_sample.cell(row=summary_sample_row, column=6, value='')
    ws_sample.cell(row=summary_sample_row, column=7, value=f'=SUM(G2:G{last_sample_data_row})')  # 数量
    ws_sample.cell(row=summary_sample_row, column=8, value=f'=SUM(H2:H{last_sample_data_row})')  # 产品成本
    ws_sample.cell(row=summary_sample_row, column=9, value=f'=SUM(I2:I{last_sample_data_row})')  # 包装成本
    ws_sample.cell(row=summary_sample_row, column=10, value=f'=SUM(J2:J{last_sample_data_row})')  # 订单总成本
    ws_sample.cell(row=summary_sample_row, column=11, value=f'=SUM(K2:K{last_sample_data_row})')  # 利润
    ws_sample.cell(row=summary_sample_row, column=12, value=f'=SUM(L2:L{last_sample_data_row})')  # 利润泰铢

    # 调整列宽 - 统一10,产品名称列18
    for col in range(1, len(sample_export_columns) + 1):
        ws_sample.column_dimensions[get_column_letter(col)].width = 10
    ws_sample.column_dimensions['D'].width = 18
    ws_sample.column_dimensions['E'].width = 18
    ws_sample.row_dimensions[1].height = 36

    # 保存文件
    wb.save(output_file)
    print(f"订单收入大表已导出: {output_file}")
    return output_file


def save_report(report, product_profit, order_details, date_folder):
    """
    保存报告到 reports 目录
    report: 汇总数据
    product_profit: 产品利润排行
    order_details: 订单明细(含利润率)
    """
    REPORTS_DIR.mkdir(exist_ok=True)

    # 保存订单分析
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = REPORTS_DIR / f"order_report_{date_folder}_{timestamp}.xlsx"

    with pd.ExcelWriter(report_file) as writer:
        # 汇总
        summary_df = pd.DataFrame([report])
        summary_df.to_excel(writer, sheet_name='汇总', index=False)

        # 订单明细(含利润率)
        order_export = order_details[[
            'Order ID', 'seller_sku', 'Product Name', 'Variation',
            'Quantity', 'Order Amount', 'income_cny',
            'product_cost_cny', 'commission_cny', 'order_cost_cny',
            'profit_cny', 'profit_rate', 'Created Time'
        ]].copy()
        order_export.columns = [
            '订单ID', '商家SKU', '产品名称', '规格',
            '数量', '订单金额(泰铢)', '收入(CNY)',
            '产品成本(CNY)', '平台佣金(CNY)', '订单总成本(CNY)',
            '利润(CNY)', '利润率', '下单时间'
        ]
        order_export = order_export.sort_values('利润(CNY)', ascending=False)
        order_export.to_excel(writer, sheet_name='订单明细', index=False)

        # 产品利润排行
        product_profit.to_excel(writer, sheet_name='产品利润排行', index=False)

    print(f"报告已保存: {report_file}")
    return report_file


# ============ 主流程 ============

def analyze(date_folder, gmv_revenue_usd=None, gmv_cost_usd=None):
    """
    主分析流程
    date_folder: 日期文件夹名,如 "20260519"
    gmv_revenue_usd: GMV广告收入(美元),需要手动输入
    gmv_cost_usd: GMV广告成本(美元),需要手动输入
    """
    print(f"开始分析 {date_folder} 的数据...")

    # 1. 加载成本表
    print("  [1/5] 加载产品成本表...")
    cost_df = load_product_cost_table()
    print(f"       成本表: {len(cost_df)} 个产品")

    # 2. 加载产品管理表
    print("  [2/5] 加载产品管理表...")
    pm_df = load_product_management()
    print(f"       产品管理表: {len(pm_df)} 个SKU")

    # 3. 加载订单数据
    print("  [3/5] 加载订单数据...")
    order_df, order_file_name = load_order_data(date_folder)
    print(f"       订单文件: {order_file_name}")
    print(f"       总订单数: {len(order_df)} 笔")

    # 4. 预处理: 去重 + 分离Normal订单和样品订单
    print("  [4/5] 预处理订单数据...")
    df_normal, df_sample = preprocess_orders(order_df)
    print(f"       Normal订单: {len(df_normal)} 笔")
    print(f"       样品订单: {len(df_sample)} 笔")

    # 5. 关联成本数据
    print("  [5/5] 关联成本数据...")
    df_normal = enrich_orders_with_cost(df_normal, cost_df, pm_df)
    df_sample = enrich_orders_with_cost(df_sample, cost_df, pm_df)

    # 6. 计算利润
    print("  [6/5] 计算利润...")
    df_normal = calculate_normal_order_profit(df_normal)
    df_sample = calculate_sample_order_profit(df_sample)

    # 7. GMV数据 - 需要手动输入(美元)
    print("  [7/5] GMV广告数据...")
    if gmv_revenue_usd is not None and gmv_cost_usd is not None:
        # 用户提供了GMV数据(美元),转换为CNY
        gmv_revenue = gmv_revenue_usd * USD_TO_CNY
        gmv_cost = gmv_cost_usd * USD_TO_CNY
        gmv_profit = gmv_revenue - gmv_cost
        gmv_roi = gmv_revenue / gmv_cost if gmv_cost > 0 else 0
        print(f"       GMV收入: {gmv_revenue_usd:.2f} USD ({gmv_revenue:.2f} CNY)")
        print(f"       GMV成本: {gmv_cost_usd:.2f} USD ({gmv_cost:.2f} CNY)")
        print(f"       GMV利润: {gmv_profit:.2f} CNY")
        print(f"       GMV ROI: {gmv_roi:.2f}")
        gmv_df = None  # 不使用表格数据
    else:
        # 尝试从表格加载
        gmv_df, gmv_file_name = load_gmv_data(date_folder)
        if not gmv_df.empty:
            gmv_df = calculate_gmv_roi(gmv_df)
            gmv_report, campaign_perf = generate_gmv_report(gmv_df)
            print(f"       GMV文件: {gmv_file_name}")
            print(f"       广告成本: {gmv_report['total_cost_cny']:.2f} CNY")
            print(f"       广告收入: {gmv_report['total_revenue_cny']:.2f} CNY")
            print(f"       平均ROI: {gmv_report['avg_roi']:.2f}")
            gmv_revenue = gmv_report['total_revenue_cny']
            gmv_cost = gmv_report['total_cost_cny']
            gmv_profit = gmv_revenue - gmv_cost
            gmv_roi = gmv_report['avg_roi']
        else:
            print("       未找到GMV数据,请手动输入")
            gmv_revenue = 0
            gmv_cost = 0
            gmv_profit = 0
            gmv_roi = 0
            gmv_df = None

    # 生成报告
    order_report, product_profit = generate_order_report(df_normal, date_folder)

    total_profit = df_normal['profit_cny'].sum() + df_sample['profit_cny'].sum()
    total_income = df_normal['income_cny'].sum() + df_sample['income_cny'].sum()
    profit_rate_str = f"{(total_profit / total_income * 100):.3f}%" if total_income > 0 else "0.000%"

    print("\n========== 订单分析报告 ==========")
    print(f"Normal订单数: {len(df_normal)}")
    print(f"样品订单数: {len(df_sample)}")
    print(f"总收入: {total_income:.3f} CNY")
    print(f"总利润: {total_profit:.3f} CNY")
    print(f"利润率: {profit_rate_str}")

    # 保存报告
    save_report(order_report, product_profit, df_normal, date_folder)

    # 导出订单收入大表 - 传入GMV数据(USD转换)
    export_order_master_table(df_normal, df_sample, date_folder, gmv_revenue, gmv_cost, gmv_roi)

    return df_normal, df_sample, gmv_df


if __name__ == "__main__":
    import sys
    import argparse

    parser = argparse.ArgumentParser(description='TikTok订单数据分析')
    parser.add_argument('date_folder', nargs='?', help='日期文件夹,如20260601')
    parser.add_argument('--gmv-revenue', type=float, help='GMV广告收入(美元)')
    parser.add_argument('--gmv-cost', type=float, help='GMV广告成本(美元)')
    args = parser.parse_args()

    if args.date_folder:
        date_folder = args.date_folder
    else:
        # 默认使用最新的日期文件夹
        date_folders = [d.name for d in DATA_DIR.iterdir() if d.is_dir() and d.name.isdigit()]
        date_folder = sorted(date_folders)[-1] if date_folders else "20260519"

    print(f"分析日期: {date_folder}")

    # 如果命令行没有提供GMV数据,尝试从表格读取
    gmv_revenue_usd = args.gmv_revenue
    gmv_cost_usd = args.gmv_cost

    df_normal, df_sample, gmv_df = analyze(date_folder, gmv_revenue_usd, gmv_cost_usd)

    print("\n分析完成!")
