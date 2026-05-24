"""
TikTok Shop Data Analytics
Main analysis script

数据处理流程:
1. 加载产品成本表 (根目录) → 提取seller_sku对应的成本信息
2. 加载订单数据 (日期文件夹) → 筛选有效订单(非取消、Normal)
3. 订单 + 产品成本表 → 通过Seller SKU关联
4. 计算利润 = 收入 - 产品成本 - 佣金 - 物流费
5. GMV广告数据 → 计算广告ROI
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font

# ============ 常量配置 ============
THB_TO_CNY = 4.7  # 泰铢兑人民币汇率
USD_TO_CNY = 7    # 美元兑人民币汇率

# 平台佣金费率
PLATFORM_COMMISSION_RATE = 0.032   # 平台佣金 3.2%
PLATFORM_GROWTH_RATE = 0.0809     # 平台增长费 8.09%
EXCHANGE_LOSS_RATE = 0.01         # 汇率损耗 1%
ORDER_FIXED_FEE_THB = 1.05       # 每笔订单固定处理费 (泰铢)

# ============ 路径配置 ============
PROJECT_ROOT = Path(__file__).parent.parent
DATA_DIR = PROJECT_ROOT / "data"
REPORTS_DIR = PROJECT_ROOT / "reports"

# ============ 数据加载 ============

def load_product_cost_table():
    """
    加载产品成本表
    data目录下的 产品成本表-XXXX.xlsx
    返回: 包含seller_sku、出厂价、海运物流单件成本、包装的DataFrame
    """
    # 查找data目录的成本表
    cost_files = list(DATA_DIR.glob("产品成本表-*.xlsx"))
    if not cost_files:
        raise FileNotFoundError("未找到产品成本表文件")

    # 使用最新的成本表
    latest_cost_file = sorted(cost_files)[-1]
    df = pd.read_excel(latest_cost_file, engine='openpyxl')

    # 提取关键列
    cost_df = df[['Seller SKU', '产品名称',
                  '出厂价',
                  '海运物流\n单件成本',
                  '包装']].copy()

    # 重命名列 - 统一为小写seller_sku便于匹配
    cost_df = cost_df.rename(columns={
        'Seller SKU': 'seller_sku',
        '产品名称': 'product_name_cn',
        '出厂价': 'factory_price',
        '海运物流\n单件成本': 'shipping_cost',
        '包装': 'package_cost'
    })

    # 转换数值类型
    for col in ['factory_price', 'shipping_cost', 'package_cost']:
        cost_df[col] = pd.to_numeric(cost_df[col], errors='coerce').fillna(0)

    # 清理: 删除重复的seller_sku和空值
    cost_df = cost_df.dropna(subset=['seller_sku'])
    cost_df = cost_df.drop_duplicates(subset=['seller_sku'], keep='first')

    return cost_df


def load_product_management():
    """
    加载产品管理表
    文件名: Tiktoksellercenter产品管理表*.xlsx
    返回: seller_sku → sku_id 映射
    """
    pm_files = list(DATA_DIR.glob("Tiktoksellercenter产品管理表*.xlsx"))
    if not pm_files:
        return pd.DataFrame()  # 返回空DF

    latest_pm_file = sorted(pm_files)[-1]
    # 跳过前两行(元数据行),使用第0行作为标题
    df = pd.read_excel(latest_pm_file, header=0, skiprows=[1, 2], engine='openpyxl')

    pm_df = df[['seller_sku', 'sku_id']].dropna(subset=['seller_sku'])
    # 去重
    pm_df = pm_df.drop_duplicates(subset=['seller_sku'], keep='first')
    return pm_df


def load_order_data(date_folder):
    """
    加载订单数据
    date_folder: 日期文件夹名,如 "20260519"
    返回: 所有订单数据(含已取消和normal、空值状态)
    """
    folder_path = DATA_DIR / date_folder
    order_files = list(folder_path.glob("*全部订单*.xlsx"))
    # 过滤掉临时文件(以~$开头)
    order_files = [f for f in order_files if not f.name.startswith('~$')]

    if not order_files:
        raise FileNotFoundError(f"未找到订单文件: {folder_path}")

    latest_order_file = sorted(order_files)[-1]
    df = pd.read_excel(latest_order_file, engine='openpyxl')

    return df, latest_order_file.name


def preprocess_orders(df):
    """
    预处理订单数据:
    1. 清理坏数据行
    2. 分离: Normal订单 vs 样品订单(空值)
    注意: 不做去重,每行代表一个订单项(同一Order ID可能有多行,代表不同商品)
    """
    # 清理坏数据行 (如 "Current order status.")
    df_clean = df[df['Order Status'] != 'Current order status.'].copy()

    # 分离Normal订单和样品订单(空值)
    df_normal = df_clean[df_clean['Normal or Pre-order'] == 'Normal'].copy()
    df_sample = df_clean[df_clean['Normal or Pre-order'].isna()].copy()

    return df_normal, df_sample


def load_gmv_data(date_folder):
    """
    加载GMV广告投放数据
    date_folder: 日期文件夹名
    返回: GMV数据DataFrame
    """
    folder_path = DATA_DIR / date_folder
    gmv_files = list(folder_path.glob("gmv_*.xlsx"))

    if not gmv_files:
        return pd.DataFrame()

    latest_gmv_file = sorted(gmv_files)[-1]
    df = pd.read_excel(latest_gmv_file, engine='openpyxl')

    return df, latest_gmv_file.name


# ============ 数据处理 ============

def enrich_orders_with_cost(order_df, cost_df, pm_df):
    """
    将订单数据与成本表关联
    通过 Seller SKU 匹配,添加:
    - 产品名称(中文)
    - 出厂价
    - 海运物流单件成本
    - 包装成本
    """
    # 统一列名大小写 - 订单表用 "Seller SKU",成本表用 "seller_sku"
    # 转换为统一格式便于匹配
    order_df = order_df.rename(columns={'Seller SKU': 'seller_sku'})

    # 合并产品成本信息
    df = order_df.merge(cost_df, on='seller_sku', how='left')

    # 合并产品管理表信息
    if not pm_df.empty:
        # 重命名pm_df的列以避免冲突
        pm_df = pm_df.rename(columns={'sku_id': 'sku_id_from_pm'})
        df = df.merge(pm_df, on='seller_sku', how='left')

    return df


def calculate_normal_order_profit(df):
    """
    计算Normal订单的利润

    收入 = 买家实付总额 (Order Amount, 泰铢) / 4.7 → 人民币

    成本项目:
    - 产品成本 = (出厂价 + 海运物流单件成本) × Quantity
    - 包装成本 = 包装费 (单件,不需要×Quantity)
    - 平台佣金 = 买家实付总额×(3.2%+8.09%+1%) + 1.05泰铢
    - 包邮服务 = 买家实付总额×7.49%

    订单成本 = 产品成本 + 包装成本 + 平台佣金 + 包邮服务

    利润 = 收入 - 订单成本
    """
    # 转换数值类型
    df['Order Amount'] = pd.to_numeric(df['Order Amount'], errors='coerce').fillna(0)

    # 收入 (泰铢 → 人民币)
    df['income_cny'] = df['Order Amount'] / THB_TO_CNY

    # 确保数值类型
    for col in ['factory_price', 'shipping_cost', 'package_cost']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)

    # 产品成本 = (出厂价 + 海运物流单件成本) × Quantity
    df['product_cost_cny'] = (
        (df['factory_price'] + df['shipping_cost']) * df['Quantity']
    )

    # 包装成本 (单件,不需要×Quantity)
    df['package_cost_cny'] = df['package_cost']

    # 平台佣金 = 买家实付总额×(3.2%+8.09%+1%) + 1.05泰铢
    commission_rate = PLATFORM_COMMISSION_RATE + PLATFORM_GROWTH_RATE + EXCHANGE_LOSS_RATE  # 12.29%
    df['commission_cny'] = (
        df['Order Amount'] * commission_rate + ORDER_FIXED_FEE_THB
    ) / THB_TO_CNY

    # 包邮服务 = 买家实付总额×7.49%
    FREE_SHIPPING_RATE = 0.0749
    df['free_shipping_cny'] = (df['Order Amount'] * FREE_SHIPPING_RATE) / THB_TO_CNY

    # 订单成本 = 产品成本 + 包装成本 + 平台佣金 + 包邮服务
    df['order_cost_cny'] = (
        df['product_cost_cny'] +
        df['package_cost_cny'] +
        df['commission_cny'] +
        df['free_shipping_cny']
    )

    # 利润 = 收入 - 订单成本
    df['profit_cny'] = df['income_cny'] - df['order_cost_cny']

    # 利润(泰铢) - 用于导出
    df['profit_thb'] = df['profit_cny'] * THB_TO_CNY

    # 利润率 (百分比)
    df['profit_rate'] = df.apply(
        lambda x: f"{(x['profit_cny'] / x['income_cny'] * 100):.3f}%" if x['income_cny'] > 0 else "0.000%",
        axis=1
    )

    # 单件利润
    df['unit_profit_cny'] = df.apply(
        lambda x: x['profit_cny'] / x['Quantity'] if x['Quantity'] > 0 else 0,
        axis=1
    )

    # 订单金额泰铢 (原始数据)
    df['order_amount_thb'] = df['Order Amount']

    return df


def calculate_sample_order_profit(df):
    """
    计算样品订单的利润

    样品订单没有收入,只有成本:
    - 产品成本 = (出厂价 + 海运物流单件成本) × Quantity
    - 包装成本 = 包装费 (单件)
    - 固定快递费 = 7元

    订单成本 = 产品成本 + 包装成本 + 固定快递费

    利润 = 0 - 订单成本 = -订单成本
    """
    # 确保数值类型
    for col in ['factory_price', 'shipping_cost', 'package_cost']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)

    # 样品没有收入
    df['income_cny'] = 0.0
    df['order_amount_thb'] = 0.0

    # 产品成本 = (出厂价 + 海运物流单件成本) × Quantity
    df['product_cost_cny'] = (
        (df['factory_price'] + df['shipping_cost']) * df['Quantity']
    )

    # 包装成本 (单件)
    df['package_cost_cny'] = df['package_cost']

    # 固定快递费 7元
    FIXED_SHIPPING_CNY = 7.0

    # 样品订单成本 = 产品成本 + 包装成本 + 固定快递费
    df['order_cost_cny'] = (
        df['product_cost_cny'] +
        df['package_cost_cny'] +
        FIXED_SHIPPING_CNY
    )

    # 样品没有平台佣金和包邮服务
    df['commission_cny'] = 0.0
    df['free_shipping_cny'] = 0.0

    # 利润 = 0 - 订单成本
    df['profit_cny'] = -df['order_cost_cny']

    # 利润(泰铢) - 用于导出 (样品利润为负)
    df['profit_thb'] = df['profit_cny'] * THB_TO_CNY

    # 利润率 (样品订单利润为负,标记为-N/A或特殊标记)
    df['profit_rate'] = '-'

    # 单件利润
    df['unit_profit_cny'] = df.apply(
        lambda x: x['profit_cny'] / x['Quantity'] if x['Quantity'] > 0 else 0,
        axis=1
    )

    return df


def calculate_gmv_roi(gmv_df):
    """
    计算GMV广告ROI
    成本和收入都是美元,转人民币计算
    """
    if gmv_df.empty:
        return gmv_df

    # 成本和收入 (美元 → 人民币)
    gmv_df['cost_cny'] = gmv_df['成本'] * USD_TO_CNY
    gmv_df['revenue_cny'] = gmv_df['总收入'] * USD_TO_CNY

    # ROI = 收入 / 成本
    gmv_df['roi'] = gmv_df.apply(
        lambda x: x['revenue_cny'] / x['cost_cny'] if x['cost_cny'] > 0 else 0,
        axis=1
    )

    # 每单成本
    gmv_df['cost_per_order_cny'] = gmv_df.apply(
        lambda x: x['cost_cny'] / x['SKU 订单数'] if x['SKU 订单数'] > 0 else 0,
        axis=1
    )

    return gmv_df


# ============ 分析报告 ============

def generate_order_report(df, date_folder):
    """
    生成订单分析报告
    """
    report = {
        'report_date': date_folder,
        'total_orders': len(df),
        'total_quantity': df['Quantity'].sum(),
        'total_income_cny': df['income_cny'].sum(),
        'total_product_cost_cny': df['product_cost_cny'].sum(),
        'total_commission_cny': df['commission_cny'].sum(),
        'total_order_cost_cny': df['order_cost_cny'].sum(),
        'total_profit_cny': df['profit_cny'].sum(),
        'avg_profit_per_order': df['profit_cny'].mean(),
        'profit_rate': df['profit_cny'].sum() / df['income_cny'].sum() if df['income_cny'].sum() > 0 else 0,
    }

    # 产品利润排行
    product_profit = df.groupby('seller_sku').agg({
        'Quantity': 'sum',
        'income_cny': 'sum',
        'product_cost_cny': 'sum',
        'commission_cny': 'sum',
        'order_cost_cny': 'sum',
        'profit_cny': 'sum',
        'Product Name': 'first',
        'factory_price': 'first',
    }).reset_index()
    product_profit = product_profit.sort_values('profit_cny', ascending=False)

    return report, product_profit


def generate_gmv_report(df):
    """
    生成GMV广告分析报告
    """
    if df.empty:
        return {'total_cost_cny': 0, 'total_revenue_cny': 0, 'total_orders': 0, 'avg_roi': 0}

    report = {
        'total_cost_cny': df['cost_cny'].sum(),
        'total_revenue_cny': df['revenue_cny'].sum(),
        'total_orders': df['SKU 订单数'].sum(),
        'avg_roi': df['roi'].mean(),
        'avg_cost_per_order': df['cost_per_order_cny'].mean(),
    }

    # 广告计划效果排行
    campaign_performance = df.groupby('广告计划名称').agg({
        'cost_cny': 'sum',
        'revenue_cny': 'sum',
        'SKU 订单数': 'sum',
        'roi': 'mean',
    }).reset_index()
    campaign_performance = campaign_performance.sort_values('roi', ascending=False)

    return report, campaign_performance


def export_order_master_table(df_normal, df_sample, date_folder):
    """
    导出订单收入大表到 data/{date}/ 目录

    包含5个Sheet:
    1. 总览 - 所有订单汇总(含全部累加行)
    2. 产品统计(Normal) - Normal订单按产品汇总
    3. 产品统计(样品) - 样品订单按产品汇总
    4. Normal订单 - 正常订单明细
    5. 样品订单 - 达人样品订单明细

    - 表头: 第1行英文, 第2行中文
    - 每行一个订单
    - 金额保留2位小数
    - 利润率显示为百分比
    - 字体: 等线, 120%大小
    """
    output_dir = DATA_DIR / date_folder
    output_dir.mkdir(exist_ok=True, parents=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"订单收入大表_{date_folder}_{timestamp}.xlsx"

    # 字体设置 - 120%大小
    font_name = '等线'
    font_size = 12  # 120% of 10

    # 定义列 (用于订单明细)
    order_columns_mapping = {
        'Order ID': '订单ID',
        'Order Status': '订单状态',
        'seller_sku': '商家SKU',
        'product_name_cn': '产品名称(中文)',
        'Product Name': '产品名称(泰语)',
        'Variation': '规格变体',
        'Quantity': '数量(件)',
        'order_amount_thb': '订单金额(泰铢)',
        'income_cny': '收入(CNY)',
        'product_cost_cny': '产品成本(CNY)',
        'package_cost_cny': '包装成本(CNY)',
        'commission_cny': '平台佣金(CNY)',
        'free_shipping_cny': '包邮服务(CNY)',
        'order_cost_cny': '订单总成本(CNY)',
        'profit_cny': '利润(CNY)',
        'profit_thb': '利润(泰铢)',
        'profit_rate': '利润率(%)',
        'Created Time': '下单时间',
    }

    def prepare_order_export_df(df):
        """准备订单导出的数据"""
        export_cols = [c for c in order_columns_mapping.keys() if c in df.columns]
        df_export = df[export_cols].copy()
        # 格式化金额 - 保留2位小数
        money_cols = ['order_amount_thb', 'income_cny', 'product_cost_cny',
                      'package_cost_cny', 'commission_cny', 'free_shipping_cny',
                      'order_cost_cny', 'profit_cny', 'profit_thb']
        for col in money_cols:
            if col in df_export.columns:
                df_export[col] = df_export[col].apply(
                    lambda x: f"{x:.2f}" if pd.notna(x) else "0.00"
                )
        df_export.columns = [order_columns_mapping.get(c, c) for c in df_export.columns]
        return df_export

    # 准备各Sheet数据
    df_all = pd.concat([df_normal, df_sample], ignore_index=True)

    # 计算汇总统计
    total_income = df_all['income_cny'].sum()
    total_cost = df_all['order_cost_cny'].sum()
    total_profit = df_all['profit_cny'].sum()
    total_income_thb = total_income * THB_TO_CNY
    total_cost_thb = total_cost * THB_TO_CNY
    total_profit_thb = total_profit * THB_TO_CNY

    normal_income = df_normal['income_cny'].sum()
    normal_cost = df_normal['order_cost_cny'].sum()
    normal_profit = df_normal['profit_cny'].sum()
    normal_income_thb = normal_income * THB_TO_CNY
    normal_cost_thb = normal_cost * THB_TO_CNY
    normal_profit_thb = normal_profit * THB_TO_CNY

    sample_income = df_sample['income_cny'].sum()
    sample_cost = df_sample['order_cost_cny'].sum()
    sample_profit = df_sample['profit_cny'].sum()
    sample_cost_thb = sample_cost * THB_TO_CNY
    sample_profit_thb = sample_profit * THB_TO_CNY

    # 计算利润率
    normal_profit_rate = (normal_profit / normal_income * 100) if normal_income > 0 else 0
    total_profit_rate = (total_profit / total_income * 100) if total_income > 0 else 0

    # 汇总统计 - 包含泰铢列
    # 顺序: Normal订单、样品订单、全部累加
    summary_data = [
        ['Normal订单', len(df_normal), f"{normal_income:.2f}", f"{normal_income_thb:.2f}",
         f"{normal_cost:.2f}", f"{normal_cost_thb:.2f}",
         f"{normal_profit:.2f}", f"{normal_profit_thb:.2f}", f"{normal_profit_rate:.2f}%"],
        ['样品订单', len(df_sample), f"{sample_income:.2f}", "0.00",
         f"{sample_cost:.2f}", f"{sample_cost_thb:.2f}",
         f"{sample_profit:.2f}", f"{sample_profit_thb:.2f}", "N/A"],
        ['全部累加', len(df_all), f"{total_income:.2f}", f"{total_income_thb:.2f}",
         f"{total_cost:.2f}", f"{total_cost_thb:.2f}",
         f"{total_profit:.2f}", f"{total_profit_thb:.2f}", f"{total_profit_rate:.2f}%"],
    ]

    def set_cell_value(ws, row, col, value):
        """设置单元格值"""
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.font = Font(name=font_name, size=font_size)
        return cell

    # 创建writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # ========== Sheet 1: 总览 ==========
        # 先创建一个空的sheet
        ws_summary = writer.book.create_sheet('总览')

        # 写入标题
        cell = ws_summary.cell(row=1, column=1, value='TikTok订单数据汇总')
        cell.font = Font(name=font_name, size=14, bold=True)

        eng_headers = ['Type', 'Order Count', 'Total Revenue(CNY)', 'Total Revenue(THB)',
                       'Total Cost(CNY)', 'Total Cost(THB)', 'Total Profit(CNY)', 'Total Profit(THB)', 'Profit Rate(%)']
        chn_headers = ['类别', '订单数', '总收入(CNY)', '总收入(泰铢)',
                       '总成本(CNY)', '总成本(泰铢)', '总利润(CNY)', '总利润(泰铢)', '利润率(%)']
        for col_idx, (eng, chn) in enumerate(zip(eng_headers, chn_headers), start=1):
            set_cell_value(ws_summary, 3, col_idx, eng)
            set_cell_value(ws_summary, 4, col_idx, chn)

        # 写入汇总数据行
        for row_idx, row_data in enumerate(summary_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                set_cell_value(ws_summary, row_idx, col_idx, value)

        # ========== Sheet 2: 产品统计(Normal) ==========
        product_stats_normal = df_normal.groupby('seller_sku').agg({
            'product_name_cn': 'first',
            'Quantity': 'sum',
            'income_cny': 'sum',
            'product_cost_cny': 'sum',
            'package_cost_cny': 'sum',
            'commission_cny': 'sum',
            'free_shipping_cny': 'sum',
            'order_cost_cny': 'sum',
            'profit_cny': 'sum',
            'order_amount_thb': 'sum',
        }).reset_index()

        product_stats_normal['unit_profit_rate'] = (
            product_stats_normal['profit_cny'] / product_stats_normal['income_cny'] * 100
            if product_stats_normal['income_cny'].sum() > 0 else 0
        )

        # 添加泰铢列
        product_stats_normal['income_thb'] = product_stats_normal['income_cny'] * THB_TO_CNY
        product_stats_normal['cost_thb'] = product_stats_normal['order_cost_cny'] * THB_TO_CNY
        product_stats_normal['profit_thb'] = product_stats_normal['profit_cny'] * THB_TO_CNY

        product_stats_normal = product_stats_normal[[
            'seller_sku', 'product_name_cn', 'Quantity',
            'income_cny', 'income_thb',
            'product_cost_cny', 'package_cost_cny', 'commission_cny', 'free_shipping_cny',
            'order_cost_cny', 'cost_thb',
            'profit_cny', 'profit_thb', 'unit_profit_rate'
        ]]

        product_stats_normal = product_stats_normal.sort_values('Quantity', ascending=False)

        product_stats_normal.columns = [
            '商家SKU', '产品名称(中文)', '订单数',
            '总收入(CNY)', '总收入(泰铢)',
            '产品成本(CNY)', '包装成本(CNY)', '平台佣金(CNY)', '包邮服务(CNY)',
            '订单总成本(CNY)', '订单总成本(泰铢)',
            '总利润(CNY)', '总利润(泰铢)', '单品利润率(%)'
        ]

        product_stats_normal.to_excel(writer, sheet_name='产品统计(Normal)', index=False, startrow=2)

        ws_product_normal = writer.sheets['产品统计(Normal)']
        eng_headers_prod = ['Seller SKU', 'Product Name(CN)', 'Order Count',
                           'Total Revenue(CNY)', 'Total Revenue(THB)',
                           'Product Cost(CNY)', 'Package Cost(CNY)', 'Commission(CNY)', 'Free Shipping(CNY)',
                           'Total Cost(CNY)', 'Total Cost(THB)',
                           'Total Profit(CNY)', 'Total Profit(THB)', 'Unit Profit Rate(%)']
        chn_headers_prod = ['商家SKU', '产品名称(中文)', '订单数',
                           '总收入(CNY)', '总收入(泰铢)',
                           '产品成本(CNY)', '包装成本(CNY)', '平台佣金(CNY)', '包邮服务(CNY)',
                           '订单总成本(CNY)', '订单总成本(泰铢)',
                           '总利润(CNY)', '总利润(泰铢)', '单品利润率(%)']
        for col_idx, (eng, chn) in enumerate(zip(eng_headers_prod, chn_headers_prod), start=1):
            set_cell_value(ws_product_normal, 1, col_idx, eng)
            set_cell_value(ws_product_normal, 2, col_idx, chn)

        # ========== Sheet 3: 产品统计(样品) ==========
        # 样品订单成本 = 产品成本 + 包装成本 + 每笔7元快递费
        FIXED_SHIPPING_CNY = 7.0

        product_stats_sample = df_sample.groupby('seller_sku').agg({
            'product_name_cn': 'first',
            'Quantity': 'sum',
            'product_cost_cny': 'sum',
            'package_cost_cny': 'sum',
            'profit_cny': 'sum',
        }).reset_index()

        # 计算快递费 (每笔订单7元)
        sample_order_counts = df_sample.groupby('seller_sku').size()
        product_stats_sample['shipping_fee_cny'] = product_stats_sample['seller_sku'].map(sample_order_counts) * FIXED_SHIPPING_CNY

        # 重新计算订单总成本
        product_stats_sample['order_cost_cny'] = (
            product_stats_sample['product_cost_cny'] +
            product_stats_sample['package_cost_cny'] +
            product_stats_sample['shipping_fee_cny']
        )

        # 重新计算利润 (无收入,利润 = -成本)
        product_stats_sample['profit_cny'] = -product_stats_sample['order_cost_cny']

        # 添加泰铢列
        product_stats_sample['cost_thb'] = product_stats_sample['order_cost_cny'] * THB_TO_CNY
        product_stats_sample['profit_thb'] = product_stats_sample['profit_cny'] * THB_TO_CNY

        product_stats_sample = product_stats_sample[[
            'seller_sku', 'product_name_cn', 'Quantity',
            'product_cost_cny', 'package_cost_cny', 'shipping_fee_cny',
            'order_cost_cny', 'cost_thb',
            'profit_cny', 'profit_thb'
        ]]

        product_stats_sample = product_stats_sample.sort_values('Quantity', ascending=False)

        product_stats_sample.columns = [
            '商家SKU', '产品名称(中文)', '订单数',
            '产品成本(CNY)', '包装成本(CNY)', '快递费(CNY)',
            '订单总成本(CNY)', '订单总成本(泰铢)',
            '总利润(CNY)', '总利润(泰铢)'
        ]

        product_stats_sample.to_excel(writer, sheet_name='产品统计(样品)', index=False, startrow=2)

        ws_product_sample = writer.sheets['产品统计(样品)']
        eng_headers_sample = ['Seller SKU', 'Product Name(CN)', 'Order Count',
                             'Product Cost(CNY)', 'Package Cost(CNY)', 'Shipping Fee(CNY)',
                             'Total Cost(CNY)', 'Total Cost(THB)',
                             'Total Profit(CNY)', 'Total Profit(THB)']
        chn_headers_sample = ['商家SKU', '产品名称(中文)', '订单数',
                             '产品成本(CNY)', '包装成本(CNY)', '快递费(CNY)',
                             '订单总成本(CNY)', '订单总成本(泰铢)',
                             '总利润(CNY)', '总利润(泰铢)']
        for col_idx, (eng, chn) in enumerate(zip(eng_headers_sample, chn_headers_sample), start=1):
            set_cell_value(ws_product_sample, 1, col_idx, eng)
            set_cell_value(ws_product_sample, 2, col_idx, chn)

        # ========== Sheet 4: Normal订单明细 ==========
        df_normal_export = prepare_order_export_df(df_normal)
        df_normal_export.to_excel(writer, sheet_name='Normal订单', index=False, startrow=2)
        ws_normal = writer.sheets['Normal订单']
        for col_idx, col_name in enumerate(df_normal_export.columns, start=1):
            set_cell_value(ws_normal, 1, col_idx, col_name)
            set_cell_value(ws_normal, 2, col_idx, col_name)

        # ========== Sheet 5: 样品订单明细 ==========
        df_sample_export = prepare_order_export_df(df_sample)
        df_sample_export.to_excel(writer, sheet_name='样品订单', index=False, startrow=2)
        ws_sample = writer.sheets['样品订单']
        for col_idx, col_name in enumerate(df_sample_export.columns, start=1):
            set_cell_value(ws_sample, 1, col_idx, col_name)
            set_cell_value(ws_sample, 2, col_idx, col_name)

        # 调整列宽
        for ws in [ws_summary, ws_product_normal, ws_product_sample, ws_normal, ws_sample]:
            for col_idx in range(1, 20):
                col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
                ws.column_dimensions[col_letter].width = 16

    print(f"订单收入大表已导出: {output_file}")
    return output_file


def save_report(report, product_profit, order_details, date_folder):
    """
    保存报告到 reports 目录
    report: 汇总数据
    product_profit: 产品利润排行
    order_details: 订单明细(含利润率)
    """
    REPORTS_DIR.mkdir(exist_ok=True)

    # 保存订单分析
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = REPORTS_DIR / f"order_report_{date_folder}_{timestamp}.xlsx"

    with pd.ExcelWriter(report_file) as writer:
        # 汇总
        summary_df = pd.DataFrame([report])
        summary_df.to_excel(writer, sheet_name='汇总', index=False)

        # 订单明细(含利润率)
        order_export = order_details[[
            'Order ID', 'seller_sku', 'Product Name', 'Variation',
            'Quantity', 'Order Amount', 'income_cny',
            'product_cost_cny', 'commission_cny', 'order_cost_cny',
            'profit_cny', 'profit_rate', 'Created Time'
        ]].copy()
        order_export.columns = [
            '订单ID', '商家SKU', '产品名称', '规格',
            '数量', '订单金额(泰铢)', '收入(CNY)',
            '产品成本(CNY)', '平台佣金(CNY)', '订单总成本(CNY)',
            '利润(CNY)', '利润率', '下单时间'
        ]
        order_export = order_export.sort_values('利润(CNY)', ascending=False)
        order_export.to_excel(writer, sheet_name='订单明细', index=False)

        # 产品利润排行
        product_profit.to_excel(writer, sheet_name='产品利润排行', index=False)

    print(f"报告已保存: {report_file}")
    return report_file


# ============ 主流程 ============

def analyze(date_folder):
    """
    主分析流程
    date_folder: 日期文件夹名,如 "20260519"
    """
    print(f"开始分析 {date_folder} 的数据...")

    # 1. 加载成本表
    print("  [1/5] 加载产品成本表...")
    cost_df = load_product_cost_table()
    print(f"       成本表: {len(cost_df)} 个产品")

    # 2. 加载产品管理表
    print("  [2/5] 加载产品管理表...")
    pm_df = load_product_management()
    print(f"       产品管理表: {len(pm_df)} 个SKU")

    # 3. 加载订单数据
    print("  [3/5] 加载订单数据...")
    order_df, order_file_name = load_order_data(date_folder)
    print(f"       订单文件: {order_file_name}")
    print(f"       总订单数: {len(order_df)} 笔")

    # 4. 预处理: 去重 + 分离Normal订单和样品订单
    print("  [4/5] 预处理订单数据...")
    df_normal, df_sample = preprocess_orders(order_df)
    print(f"       Normal订单: {len(df_normal)} 笔")
    print(f"       样品订单: {len(df_sample)} 笔")

    # 5. 关联成本数据
    print("  [5/5] 关联成本数据...")
    df_normal = enrich_orders_with_cost(df_normal, cost_df, pm_df)
    df_sample = enrich_orders_with_cost(df_sample, cost_df, pm_df)

    # 6. 计算利润
    print("  [6/5] 计算利润...")
    df_normal = calculate_normal_order_profit(df_normal)
    df_sample = calculate_sample_order_profit(df_sample)

    # 7. 加载GMV数据
    print("  [7/5] 加载并分析GMV广告数据...")
    gmv_df, gmv_file_name = load_gmv_data(date_folder)
    if not gmv_df.empty:
        gmv_df = calculate_gmv_roi(gmv_df)
        gmv_report, campaign_perf = generate_gmv_report(gmv_df)
        print(f"       GMV文件: {gmv_file_name}")
        print(f"       广告成本: {gmv_report['total_cost_cny']:.2f} CNY")
        print(f"       广告收入: {gmv_report['total_revenue_cny']:.2f} CNY")
        print(f"       平均ROI: {gmv_report['avg_roi']:.2f}")
    else:
        print("       未找到GMV数据")

    # 生成报告
    order_report, product_profit = generate_order_report(df_normal, date_folder)

    total_profit = df_normal['profit_cny'].sum() + df_sample['profit_cny'].sum()
    total_income = df_normal['income_cny'].sum() + df_sample['income_cny'].sum()
    profit_rate_str = f"{(total_profit / total_income * 100):.3f}%" if total_income > 0 else "0.000%"

    print("\n========== 订单分析报告 ==========")
    print(f"Normal订单数: {len(df_normal)}")
    print(f"样品订单数: {len(df_sample)}")
    print(f"总收入: {total_income:.3f} CNY")
    print(f"总利润: {total_profit:.3f} CNY")
    print(f"利润率: {profit_rate_str}")

    # 保存报告
    save_report(order_report, product_profit, df_normal, date_folder)

    # 导出订单收入大表
    export_order_master_table(df_normal, df_sample, date_folder)

    return df_normal, df_sample, gmv_df


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        date_folder = sys.argv[1]
    else:
        # 默认使用最新的日期文件夹
        date_folders = [d.name for d in DATA_DIR.iterdir() if d.is_dir() and d.name.isdigit()]
        date_folder = sorted(date_folders)[-1] if date_folders else "20260519"

    print(f"分析日期: {date_folder}")

    df_normal, df_sample, gmv_df = analyze(date_folder)

    print("\n分析完成!")